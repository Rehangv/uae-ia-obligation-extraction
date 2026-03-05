
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"   # deep navy      — header row background
CLR_HEADER_FG   = "FFFFFF"   # white          — header row text
CLR_SUBHDR_BG   = "2E75B6"   # medium blue    — section group rows
CLR_SUBHDR_FG   = "FFFFFF"
CLR_ROW_ALT     = "EDF2F9"   # very light blue — alternating data rows
CLR_ROW_NORM    = "FFFFFF"   # white           — normal data rows

# Gap analysis status colours (subtle, not harsh)
CLR_NO_CHANGE   = "FFFFFF"   # white
CLR_CHANGED     = "FFF3CD"   # soft amber
CLR_NEW         = "E8F5E9"   # soft green
CLR_REMOVED     = "FDECEA"   # soft red/pink
CLR_ERROR       = "F5F5F5"   # light grey

STATUS_COLOURS = {
    "No Change":   CLR_NO_CHANGE,
    "Changed":     CLR_CHANGED,
    "New Section": CLR_NEW,
    "Removed":     CLR_REMOVED,
    "Error":       CLR_ERROR,
}


# ── Reusable style builders ───────────────────────────────────────────────────

def header_font(size=10):
    return Font(name="Calibri", bold=True, color=CLR_HEADER_FG, size=size)

def body_font(size=9, bold=False):
    return Font(name="Calibri", size=size, bold=bold)

def header_fill():
    return PatternFill("solid", start_color=CLR_HEADER_BG)

def status_fill(status: str):
    colour = STATUS_COLOURS.get(status, CLR_ROW_NORM)
    return PatternFill("solid", start_color=colour)

def alt_fill(row_idx: int):
    colour = CLR_ROW_ALT if row_idx % 2 == 0 else CLR_ROW_NORM
    return PatternFill("solid", start_color=colour)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def top_align(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row_num: int, n_cols: int, height: int = 35):
    """Apply deep navy header styling to a row."""
    ws.row_dimensions[row_num].height = height
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = header_font()
        cell.fill      = header_fill()
        cell.alignment = center_align()
        cell.border    = thin_border()


def style_data_row(ws, row_num: int, n_cols: int,
                   height: int = 80,
                   fill: PatternFill = None,
                   alt: bool = False):
    """Apply body styling to a data row."""
    ws.row_dimensions[row_num].height = height
    row_fill = fill or alt_fill(row_num)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = body_font()
        cell.fill      = row_fill
        cell.alignment = top_align()
        cell.border    = thin_border()


def freeze_and_filter(ws, freeze_cell: str = "B2"):
    """Freeze top row and enable auto-filter."""
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions


def set_col_widths(ws, widths: list):
    """Set column widths from a list."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_summary_sheet(ws):
    """Style the summary sheet with the same palette."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12

    # Title
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=CLR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 28

    # Blank row
    ws.row_dimensions[2].height = 8

    # Column headers
    for col, val in [("A", "Status"), ("B", "Count")]:
        cell = ws[f"{col}3"]
        cell.value     = val
        cell.font      = header_font(size=9)
        cell.fill      = header_fill()
        cell.alignment = center_align()
        cell.border    = thin_border()
    ws.row_dimensions[3].height = 22

    # Data rows with status colours
    for row_num in range(4, ws.max_row + 1):
        status_cell = ws.cell(row=row_num, column=1)
        count_cell  = ws.cell(row=row_num, column=2)
        if not status_cell.value:
            continue
        status = str(status_cell.value)
        fill   = PatternFill("solid", start_color=STATUS_COLOURS.get(status, CLR_ROW_NORM))
        for cell in [status_cell, count_cell]:
            cell.font      = body_font(size=9)
            cell.fill      = fill
            cell.alignment = center_align()
            cell.border    = thin_border()
        ws.row_dimensions[row_num].height = 20

    # Legend
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = Font(
        name="Calibri", bold=True, size=9, color=CLR_HEADER_BG)
    ws.row_dimensions[legend_row].height = 16

    legend_items = [
        ("No Change",   CLR_NO_CHANGE,  "Obligations identical in both versions"),
        ("Changed",     CLR_CHANGED,    "Obligations materially changed in PDF2"),
        ("New Section", CLR_NEW,        "Section only exists in PDF2"),
        ("Removed",     CLR_REMOVED,    "Section only exists in PDF1"),
    ]
    for i, (label, colour, desc) in enumerate(legend_items, start=legend_row + 1):
        lbl  = ws.cell(row=i, column=1, value=label)
        desc_cell = ws.cell(row=i, column=2, value=desc)
        lbl.font       = body_font(size=8, bold=True)
        lbl.fill       = PatternFill("solid", start_color=colour)
        lbl.alignment  = center_align(wrap=False)
        lbl.border     = thin_border()
        desc_cell.font      = body_font(size=8)
        desc_cell.alignment = top_align(wrap=False)
        ws.row_dimensions[i].height = 16
        ws.column_dimensions["B"].width = 45