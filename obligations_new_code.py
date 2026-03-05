import os
import re
import json
import sys
import time
import asyncio
import base64
from typing import List

# -------------------- Networking ---------------------------
import requests
import websockets
from websockets.sync.client import connect

# -------------------- Data & Utilities ---------------------
import pandas as pd
from tqdm import tqdm
from fuzzywuzzy import fuzz
from asgiref.sync import sync_to_async

# -------------------- Excel Handling -----------------------
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# -------------------- PDF Handling -------------------------
from PyPDF2 import PdfReader

# -------------------- Project Constants --------------------
from status_code import *
from rbi_constants import *

# -------------------- LangChain (Gemini ONLY) --------------
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.prompts import PromptTemplate

# Google API Key Pool
GOOGLE_API_KEYS = [   
    "",  # Api key 1
    "",  # Api key 2
    ""   # Api key 3
]
_current_key_index = 0   

def _get_next_key() -> str:
    """Rotate to next key in pool and return it."""
    global _current_key_index
    _current_key_index = (_current_key_index + 1) % len(GOOGLE_API_KEYS)
    return GOOGLE_API_KEYS[_current_key_index]

os.environ["GOOGLE_API_KEY"] = GOOGLE_API_KEYS[0]



# >>>>>>>>>>>>>>>>> >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Hyper link extraction >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

print("Obligation code")
def update_no_action(obligation):
    apology_phrases = ['sorry', 'apologise', 'apologize', 'apologies', 'chatbot', 'no actions', 'no action', 'no actions.']
    
    # Convert obligation to lowercase for case-insensitive matching
    obligation_lower = obligation.lower()

    # Check if any apology phrase is in the obligation
    if any(phrase in obligation_lower for phrase in apology_phrases):
        return "No actions.", True
    
    return obligation, False


def classify_obligation(obligation):
    try:
        # Convert to lowercase and clean the text
        obligation = obligation.lower()

        # Regex to find the classification
        classification_pattern = r'classification:\s*(filing/return|display|procedural requirement)'
        classifications = re.findall(classification_pattern, obligation)

        # Capitalize the classifications found by regex
        if classifications:
            return ', '.join(set(classification.capitalize() for classification in classifications))

        # Check for keywords in the raw text
        keywords = ['procedural requirement', 'filing/return', 'display']
        found_keywords = [keyword.capitalize() for keyword in keywords if keyword in obligation]

        if found_keywords:
            return ', '.join(found_keywords)

        return 'N/A'

    except Exception as e:
        return f"Error occurred: {str(e)}"

##############
def read_json(json_fl_path):
    try:
        with open(json_fl_path, 'r', encoding='utf-8') as file:
            data = json.load(file)  # Parse the JSON data
        return data  # Return parsed JSON data
    except UnicodeDecodeError as e:
        print(f"Unicode decode error occurred while reading {json_fl_path}: {e}")
        return None  # Return None or handle the error appropriately
    except json.JSONDecodeError as e:
        print(f"JSON decode error occurred in {json_fl_path}: {e}")
        return None  # Return None or handle the error appropriately
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None  # Return None or handle the error appropriately



def extract_page_reference(hyperlink):
    match = re.search(r'(Page_\d+(_\d+)?)', hyperlink)
    if match:
        return match.group(1)
    return None


def table_obl(output_directory):
    excel_file_path = os.path.join(output_directory, "Working_outputss.xlsx")
    json_fl_path = os.path.join(output_directory, "tab.json")
    df = pd.read_excel(excel_file_path, sheet_name='Sheet1') # can be replace
    data = read_json(json_fl_path)
    
    for index, row in df.iterrows(): # need to pass index val
        get_pagenumber = row['Page Number']
        if pd.notna(get_pagenumber):
            page_reference = extract_page_reference(get_pagenumber)
            if page_reference:
                match = re.match(r'Page_(\d+)(_(\d+))?', page_reference)
                if match:
                    page_number = int(match.group(1))
                    item_index = int(match.group(3)) - 1 if match.group(3) else 0  # Default to 0 if no index
                    page_data = data[page_number]
                    table_val = page_data.get(str(page_number))[item_index][1]
                    print(f"Extracted item from page {page_number}: {table_val}")
                else:
                    print(f"Page number {page_number} out of range in data")


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.. Nagotiation Api ?>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

def test_negotiation_api():
    payload={
    "temperature": "0.1",
    "top_p": "0.95",
    "max_tokens": "500",
    "presence_penalty": "0",
    "frequency_penalty": "0",
    "model_name": "YOUR_MODEL_NAME_HERE",
    }
    headers = {
        'Content-Type': 'application/json',
        "X_API_KEY": "YOUR_API_KEY_HERE"
        # 'Authorization': 'Bearer YOUR_ACCESS_TOKEN',  # Uncomment if you need to add an Authorization header
    }




    negotiate_data_url = "YOUR_NEGOTIATION_API_URL_HERE"
    # negotiate_data_url="YOUR_NEGOTIATION_API_URL_HERE"
    # response=rq.post(url=negotiate_data_url,data=json.dumps(payload))
    response = requests.post(negotiate_data_url, json=payload, headers=headers)
    response.raise_for_status() 
    
    if response.status_code == 200:
        response_json=response.json()
        return response_json['url'],response_json['token']
        
    else:
        print("response_jeson: ", str(response.json()))
        print("Request to negotiate api failed")
        sys.exit(1)


def get_obligation_api1(paragraph, semaphore):
  
    if paragraph not in ("No actions.", "No actions", "No action","No action."):
        # delay_time = 10
        # time.sleep(delay_time)
        llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-flash-lite",          # Latest stable Gemini Flash – very fast & capable          # Free tier, fast & capable
        temperature=0.0,                    # Deterministic output for obligations
        max_output_tokens=2048,             # Enough for long responses
        google_api_key=os.environ["GOOGLE_API_KEY"]
        )


        prompt_template = f"""
        {{system_message}}
        Now, answer the following question based on the information provided:
        Question: {{question}}
        """

        prompt = PromptTemplate(template=prompt_template, input_variables=["system_message", "question"])

        responses = []
    
        
        system_message = create_system_message2(paragraph)

        final_prompt = prompt.format(
        system_message=system_message, 
        question="""Analyse the provided text to extract the following compliance information in sequential order:
        1. Compliance Classification: Determine the overall category of the compliance obligation. If you can't find anything, write 'NA'.
        2. Sub Classification: Identify the specific subcategory within the Compliance Classification. If you can't find anything, write 'NA'.
        3. Compliance Header: Extract the main heading or title of the compliance obligation.
        4. Penalty Type: If applicable, determine the type of penalty associated with non-compliance (e.g., fine, imprisonment, license revocation).
        5. Compliance Type: Identify whether the compliance obligation is event-based, ongoing, or one-time.
        6. Event Name: If the compliance obligation is event-based, extract the specific event that triggers the obligation.
        7. Event Applicability: Determine the timeframe or conditions under which the event-based obligation must be fulfilled.
        8. Event Due Date: If applicable, identify the specific deadline or due date for fulfilling the obligation.
        Please provide your answer in the same format as the example below.
        Example:
        Text: "The bank must report suspicious activity to the regulatory authority within 24 hours of detection."
        Output:
        Compliance Classification: Financial Crime
        Sub Classification: Anti-Money Laundering
        Compliance Header: Suspicious Activity Reporting
        Penalty Type: Fine
        Compliance Type: Event-based
        Event Name: Detection of suspicious activity
        Event Applicability: Within 24 hours
        Event Due Date: X"""
        )

   
            
        messages = [
            {"role": "system", "content": system_message},
            {"role": "user", "content": final_prompt}
        ]
            
        response = llm.invoke(messages)
        responses.append(response.content)
    
    else:
        return "No Action."
    
    return responses



async def get_obligation_apiii111(paragraph :str,access_token: str,semaphore) -> str:
    async with semaphore:
        if paragraph not in ("No actions.", "No actions", "No action","No action."):
            data=[]
            #import pdb; pdb.set_trace()
            #url = "YOUR_WEBSOCKET_API_URL_HERE"
            url = "YOUR_WEBSOCKET_API_URL_HERE"
            my_dict = {}
            async with websockets.connect(url,timeout=50) as ws:
                out_message_dict={
                    
                            "ce-type": "user_message",
                            "query": [ 
                                {
                                    "role": "system",
                                    "content": """
                                        You are AI chatbot who is expert in regulatory compliance for banks and other financial institutions.
                                        1. Analyse regulatory circulars/text to determine if there are penalties for banks that don't comply. If there's a fine, write 'Fine' in the 'Penalty Type' column. If there is imprisonment, write 'Imprisonment.' If there is license revocation, write 'License Revoke.' For other penalties, write 'Other.' If you can't find anything, write 'NA'.
                                        2. If there are actions/obligations for banks classify the ‘Compliance Classification’ and ‘Sub Classification’ as one of the following:
                                        a.     If the text mentions a bank's obligation on the following topics related to ‘Governance’ -BoD/Secretarial, Control functions, Corporate Governance, Corporate Social Responsibility, Licensing, MOA/AOA, Secretarial, Financial services and Cheques, then the Compliance Classification is 'Governance' and the Sub Classification is the specific topic associated with the obligation.
                                        b.    If the text mentions a bank's obligation on the following topics related to 'Account opening (including KYC)' - Customer identification, DED, Due diligence, Fees and charges and Corporate Agency, then the Compliance Classification is 'Account opening (including KYC)' and the Sub Classification is the specific topic associated with the obligation.
                                        c.    If the text mentions a bank's obligation on the following topics related to 'Bancatakaful' - Administrative, then the Compliance Classification is 'Bancatakaful' and the Sub Classification is the specific topic associated with the obligation.
                                        d.   If the text mentions a bank's obligation on the following topics related to 'Branch banking' - ATMs, Cash, Cheques, Customer forex, Licensing, Opening, closing and merging of branches/ATMs, Reporting requirements (including BRFs), Traveller’s cheques and FATCA/CRS, then the Compliance Classification is 'Branch banking' and the Sub Classification is the specific topic associated with the obligation.
                                        e.   If the text mentions a bank's obligation on the following topics related to 'Compliance' - Regulatory affairs and Credit appraisal, then the Compliance Classification is 'Compliance' and the Sub Classification is the specific topic associated with the obligation.
                                        f.    If the text mentions a bank's obligation on the following topics related to 'Credit Risk' - Account balances, then the Compliance Classification is 'Credit Risk' and the Sub Classification is the specific topic associated with the obligation.
                                        g.    If the text mentions a bank's obligation on the following topics related to 'Customer Service & Fair Lending Practices' - Account closure, ATMs, Complaints and grievances, Credit appraisal, Customer confidentiality, Dormant Accounts, Fees and charges, Inclusivity and diversity, Scheme and Sukuk, then the Compliance Classification is 'Customer Service & Fair Lending Practices' and the Sub Classification is the specific topic associated with the obligation.
                                        h.   If the text mentions a bank's obligation on the following topics related to 'Debt' - Accounting, then the Compliance Classification is 'Debt' and the Sub Classification is the specific topic associated with the obligation.
                                        i.     If the text mentions a bank's obligation on the following topics related to 'Finance' - Balance of Payments, Capital Adequacy Framework, IFRS, Liquidity risk, Reporting requirements (including BRFs), Investments and Share capital, then the Compliance Classification is 'Finance' and the Sub Classification is the specific topic associated with the obligation.
                                        j.   If the text mentions a bank's obligation on the following topics related to 'Financial Crime' - Anti-Money Laundering, Reporting requirements (including BRFs), Sanctions, Seizures and Accounting, then the Compliance Classification is 'Financial Crime' and the Sub Classification is the specific topic associated with the obligation.
                                        k.   If the text mentions a bank's obligation on the following topics related to 'General' - Administrative, Dormant Accounts, Fees and charges, Regulatory affairs, Training & Awareness, then the Compliance Classification is 'General' and the Sub Classification is the specific topic associated with the obligation.
                                        l.    If the text mentions a bank's obligation on the following topics related to 'Human Resources'  - Emiratisation, Employee separation, Reporting requirements (including BRFs) and Control framework, then the Compliance Classification is 'Human Resources' and the Sub Classification is the specific topic associated with the obligation.
                                        m.  If the text mentions a bank's obligation on the following topics related to 'Information & Cyber Security' - NESA, Certificate of deposit, then the Compliance Classification is 'Information & Cyber Security' and the Sub Classification is the specific topic associated with the obligation.
                                        n.     If the text mentions a bank's obligation on the following topics related to 'Investor Services' - Financial Free Zones, Investments, Reporting requirements (including BRFs), Share capital, Disclosure requirements, Dividends, Investments, Structured Products, Trading window and Compensation Structure, then the Compliance Classification is 'Investor Services' and the Sub Classification is the specific topic associated with the obligation.
                                        o.   If the text mentions a bank's obligation on the following topics related to 'IT infrastructure'  - Reporting requirements (including BRFs), then the Compliance Classification is 'IT infrastructure' and the Sub Classification is the specific topic associated with the obligation.
                                        p.   If the text mentions a bank's obligation on the following topics related to 'Legal' - Arbitration, Court cases, then the Compliance Classification is 'Legal' and the Sub Classification is the specific topic associated with the obligation.
                                        q.    If the text mentions a bank's obligation on the following topics related to 'Loans/lending' -  Balance transfer, Collateral management, Corporate loans, Credit appraisal, Fees and charges, Loan Against Gold, Loan Against Property, Real estate finance, Repayment, Reporting requirements (including BRFs), Retail finance and Transfer of loans, then the Compliance Classification is 'Loans/lending' and the Sub Classification is the specific topic associated with the obligation.
                                        r.     If the text mentions a bank's obligation on the following topics related to 'Payments & Remittances' - Cards, Control framework, DDS, Digital banking, Due diligence, Frauds and fund recall, FTS, ICCS, Inward remittances, Reporting requirements (including BRFs), SWIFT payments and SWITCH, then the Compliance Classification is 'Payments & Remittances' and the Sub Classification is the specific topic associated with the obligation.
                                        s.    If the text mentions a bank's obligation on the following topics related to 'Premises & Admin'  - Buildings and surrounding areas and Security, then the Compliance Classification is 'Premises & Admin' and the Sub Classification is the specific topic associated with the obligation.
                                        t.    If the text mentions a bank's obligation on the following topics related to 'Prudential Norms, Statutory Ratios & Risk Limits' - Advances to deposit ratio, Cash Reserve Requirements and Credit concentration, then the Compliance Classification is 'Prudential Norms, Statutory Ratios & Risk Limits' and the Sub Classification is the specific topic associated with the obligation.
                                        u.   If the text mentions a bank's obligation on the following topics related to 'Risk Management'  - Governance, Country Risk, Fraud risk, Interest Rate Risk, Liquidity risk, Market Risk, Inter Bank Dealings and Operational Risk, then the Compliance Classification is 'Risk Management' and the Sub Classification is the specific topic associated with the obligation.
                                        v.    If the text mentions a bank's obligation on the following topics related to 'Shari'a compliance' - Certificate of deposit, Early settlement of loans, HSA, Products, Sale of debt, Sukuk and Zakat, then the Compliance Classification is 'Shari'a compliance' and the Sub Classification is the specific topic associated with the obligation.
                                        w.   If the text mentions a bank's obligation on the following topics related to 'Trade Finance'  - Exports and Imports, then the Compliance Classification is 'Trade Finance' and the Sub Classification is the specific topic associated with the obligation.
                                        x.    If the text mentions a bank's obligation on the following topics related to 'Treasury' -  Certificate of deposit, Customer forex, Derivatives, Reporting requirements (including BRFs), Call and Notice Money market, then the Compliance Classification is 'Treasury' and the Sub Classification is the specific topic associated with the obligation.

                                        3. If there are actions/obligations for banks then ‘Compliance Header’ will be 3-6 words header from the actions/obligations. 
                                        Example: 
                                                            Bank’s obligation: Establish proper procedures and systems and provide necessary instructions to all relevant bank or branch staff to ensure compliance with these guidelines.
                                                            Compliance Header: Systems to educate bank staff for these guidelines
                                        4. Classify the compliance type as 'One-time', 'Ongoing' or 'Event-based'.
                                        a. One-time Compliances: Compliance actions that are required only once, typically tied to a specific event, regulatory change, or business action, e.g., applying for licenses, initial regulatory setups, or entity formation. Note: There are almost no chances that One-time compliance will repeat.
                                        b. Ongoing Compliances: Requires continuous or recurring actions for the bank. Part of regular business operations, e.g., AML monitoring, KYC updates, regular reporting, Financial reporting, risk management, reason for rejection and compliance training.
                                        c. Event-based Compliance: Compliance actions that are triggered by specific events, e.g. In case of rejection of any Consumers' application for a Financial Product and/or Service by the Licensed Financial Institution, the institution must disclose the reason for rejection to the applicant, except where the reason for rejection is related to Financial Crime Compliance risks or as may be prohibited by law.
                                        5. Event Name: If the text mentions a bank's obligation and the compliance type is 'Event-based,' identify the specific event described in the text as Event Name. 
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                    Event Name: Reporting to RBI of an Account red flagged on the CRILC platform reporting threshold by the reporting entity.
                                        6. Event Applicability: If the text mentions a bank's obligation and the compliance type is 'Event-based,' describe the specific circumstances or conditions under which a compliance obligation must be fulfilled. It often ties the obligation to a particular event or occurrence.
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                    Event Applicability: Within 7 days of account being red flagged
                                        7. Event - Due Date: If the text mentions a bank's obligation and the compliance type is 'Event-based,' identify any due date, time or description available in the text as Event – Due Date. If you can't find anything, write 'X'.
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                Even Due Date: X+7 days

                                    """ + str(paragraph)
                                },
                                {
                                    "role": "user",
                                    "content": """
                                    Analyse the provided text to extract the following compliance information in sequential order:
                                    1.Compliance Classification: Determine the overall category of the compliance obligation. If you can't find anything, write 'NA'.
                                    2.Sub Classification: Identify the specific subcategory within the Compliance Classification. If you can't find anything, write 'NA'.
                                    3.Compliance Header: Extract the main heading or title of the compliance obligation.
                                    4.Penalty Type: If applicable, determine the type of penalty associated with non-compliance (e.g., fine, imprisonment, license revocation).
                                    5.Compliance Type: Identify whether the compliance obligation is event-based, ongoing, or one-time.
                                    6.Event Name: If the compliance obligation is event-based, extract the specific event that triggers the obligation.
                                    7.Event Applicability: Determine the timeframe or conditions under which the event-based obligation must be fulfilled.
                                    8.Event Due Date: If applicable, identify the specific deadline or due date for fulfilling the obligation.
                                    Please provide your answer in the same format as the example below.
                                    Example:
                                    Text: "The bank must report suspicious activity to the regulatory authority within 24 hours of detection."
                                    Output:
                                    Compliance Classification: Financial Crime
                                    Sub Classification: Anti-Money Laundering
                                    Compliance Header: Suspicious Activity Reporting
                                    Penalty Type: Fine
                                    Compliance Type: Event-based
                                    Event Name: Detection of suspicious activity
                                    Event Applicability: Within 24 hours
                                    Event Due Date: X

                                    """
                                }
                            ],
                            "access_token": access_token
                            }

                out_message = json.dumps(out_message_dict)
                await ws.send(out_message)

                data = ""
                while True:
                    recv_message = await asyncio.wait_for(ws.recv(), timeout=20)
                    # Parse the received JSON message and extract the data
                    message_json = json.loads(recv_message)

                    if message_json.get("data", "") == "verbose end" or message_json.get("data", "") == "stream end":
                        break
                    data += message_json.get("data", "")

                return data.replace('"', '')
        else:
            'No actions.'


def create_system_message2(excel_column_data):
    system_message = f"""
        You are AI chatbot who is expert in regulatory compliance for banks and other financial institutions.
                                        1. Analyse regulatory circulars/text to determine if there are penalties for banks that don't comply. If there's a fine, write 'Fine' in the 'Penalty Type' column. If there is imprisonment, write 'Imprisonment.' If there is license revocation, write 'License Revoke.' For other penalties, write 'Other.' If you can't find anything, write 'NA'.
                                        2. If there are actions/obligations for banks classify the ‘Compliance Classification’ and ‘Sub Classification’ as one of the following:
                                        a.     If the text mentions a bank's obligation on the following topics related to ‘Governance’ -BoD/Secretarial, Control functions, Corporate Governance, Corporate Social Responsibility, Licensing, MOA/AOA, Secretarial, Financial services and Cheques, then the Compliance Classification is 'Governance' and the Sub Classification is the specific topic associated with the obligation.
                                        b.    If the text mentions a bank's obligation on the following topics related to 'Account opening (including KYC)' - Customer identification, DED, Due diligence, Fees and charges and Corporate Agency, then the Compliance Classification is 'Account opening (including KYC)' and the Sub Classification is the specific topic associated with the obligation.
                                        c.    If the text mentions a bank's obligation on the following topics related to 'Bancatakaful' - Administrative, then the Compliance Classification is 'Bancatakaful' and the Sub Classification is the specific topic associated with the obligation.
                                        d.   If the text mentions a bank's obligation on the following topics related to 'Branch banking' - ATMs, Cash, Cheques, Customer forex, Licensing, Opening, closing and merging of branches/ATMs, Reporting requirements (including BRFs), Traveller’s cheques and FATCA/CRS, then the Compliance Classification is 'Branch banking' and the Sub Classification is the specific topic associated with the obligation.
                                        e.   If the text mentions a bank's obligation on the following topics related to 'Compliance' - Regulatory affairs and Credit appraisal, then the Compliance Classification is 'Compliance' and the Sub Classification is the specific topic associated with the obligation.
                                        f.    If the text mentions a bank's obligation on the following topics related to 'Credit Risk' - Account balances, then the Compliance Classification is 'Credit Risk' and the Sub Classification is the specific topic associated with the obligation.
                                        g.    If the text mentions a bank's obligation on the following topics related to 'Customer Service & Fair Lending Practices' - Account closure, ATMs, Complaints and grievances, Credit appraisal, Customer confidentiality, Dormant Accounts, Fees and charges, Inclusivity and diversity, Scheme and Sukuk, then the Compliance Classification is 'Customer Service & Fair Lending Practices' and the Sub Classification is the specific topic associated with the obligation.
                                        h.   If the text mentions a bank's obligation on the following topics related to 'Debt' - Accounting, then the Compliance Classification is 'Debt' and the Sub Classification is the specific topic associated with the obligation.
                                        i.     If the text mentions a bank's obligation on the following topics related to 'Finance' - Balance of Payments, Capital Adequacy Framework, IFRS, Liquidity risk, Reporting requirements (including BRFs), Investments and Share capital, then the Compliance Classification is 'Finance' and the Sub Classification is the specific topic associated with the obligation.
                                        j.   If the text mentions a bank's obligation on the following topics related to 'Financial Crime' - Anti-Money Laundering, Reporting requirements (including BRFs), Sanctions, Seizures and Accounting, then the Compliance Classification is 'Financial Crime' and the Sub Classification is the specific topic associated with the obligation.
                                        k.   If the text mentions a bank's obligation on the following topics related to 'General' - Administrative, Dormant Accounts, Fees and charges, Regulatory affairs, Training & Awareness, then the Compliance Classification is 'General' and the Sub Classification is the specific topic associated with the obligation.
                                        l.    If the text mentions a bank's obligation on the following topics related to 'Human Resources'  - Emiratisation, Employee separation, Reporting requirements (including BRFs) and Control framework, then the Compliance Classification is 'Human Resources' and the Sub Classification is the specific topic associated with the obligation.
                                        m.  If the text mentions a bank's obligation on the following topics related to 'Information & Cyber Security' - NESA, Certificate of deposit, then the Compliance Classification is 'Information & Cyber Security' and the Sub Classification is the specific topic associated with the obligation.
                                        n.     If the text mentions a bank's obligation on the following topics related to 'Investor Services' - Financial Free Zones, Investments, Reporting requirements (including BRFs), Share capital, Disclosure requirements, Dividends, Investments, Structured Products, Trading window and Compensation Structure, then the Compliance Classification is 'Investor Services' and the Sub Classification is the specific topic associated with the obligation.
                                        o.   If the text mentions a bank's obligation on the following topics related to 'IT infrastructure'  - Reporting requirements (including BRFs), then the Compliance Classification is 'IT infrastructure' and the Sub Classification is the specific topic associated with the obligation.
                                        p.   If the text mentions a bank's obligation on the following topics related to 'Legal' - Arbitration, Court cases, then the Compliance Classification is 'Legal' and the Sub Classification is the specific topic associated with the obligation.
                                        q.    If the text mentions a bank's obligation on the following topics related to 'Loans/lending' -  Balance transfer, Collateral management, Corporate loans, Credit appraisal, Fees and charges, Loan Against Gold, Loan Against Property, Real estate finance, Repayment, Reporting requirements (including BRFs), Retail finance and Transfer of loans, then the Compliance Classification is 'Loans/lending' and the Sub Classification is the specific topic associated with the obligation.
                                        r.     If the text mentions a bank's obligation on the following topics related to 'Payments & Remittances' - Cards, Control framework, DDS, Digital banking, Due diligence, Frauds and fund recall, FTS, ICCS, Inward remittances, Reporting requirements (including BRFs), SWIFT payments and SWITCH, then the Compliance Classification is 'Payments & Remittances' and the Sub Classification is the specific topic associated with the obligation.
                                        s.    If the text mentions a bank's obligation on the following topics related to 'Premises & Admin'  - Buildings and surrounding areas and Security, then the Compliance Classification is 'Premises & Admin' and the Sub Classification is the specific topic associated with the obligation.
                                        t.    If the text mentions a bank's obligation on the following topics related to 'Prudential Norms, Statutory Ratios & Risk Limits' - Advances to deposit ratio, Cash Reserve Requirements and Credit concentration, then the Compliance Classification is 'Prudential Norms, Statutory Ratios & Risk Limits' and the Sub Classification is the specific topic associated with the obligation.
                                        u.   If the text mentions a bank's obligation on the following topics related to 'Risk Management'  - Governance, Country Risk, Fraud risk, Interest Rate Risk, Liquidity risk, Market Risk, Inter Bank Dealings and Operational Risk, then the Compliance Classification is 'Risk Management' and the Sub Classification is the specific topic associated with the obligation.
                                        v.    If the text mentions a bank's obligation on the following topics related to 'Shari'a compliance' - Certificate of deposit, Early settlement of loans, HSA, Products, Sale of debt, Sukuk and Zakat, then the Compliance Classification is 'Shari'a compliance' and the Sub Classification is the specific topic associated with the obligation.
                                        w.   If the text mentions a bank's obligation on the following topics related to 'Trade Finance'  - Exports and Imports, then the Compliance Classification is 'Trade Finance' and the Sub Classification is the specific topic associated with the obligation.
                                        x.    If the text mentions a bank's obligation on the following topics related to 'Treasury' -  Certificate of deposit, Customer forex, Derivatives, Reporting requirements (including BRFs), Call and Notice Money market, then the Compliance Classification is 'Treasury' and the Sub Classification is the specific topic associated with the obligation.

                                        3. If there are actions/obligations for banks then ‘Compliance Header’ will be 3-6 words header from the actions/obligations. 
                                        Example: 
                                                            Bank’s obligation: Establish proper procedures and systems and provide necessary instructions to all relevant bank or branch staff to ensure compliance with these guidelines.
                                                            Compliance Header: Systems to educate bank staff for these guidelines
                                        4. Classify the compliance type as 'One-time', 'Ongoing' or 'Event-based'.
                                        a. One-time Compliances: Compliance actions that are required only once, typically tied to a specific event, regulatory change, or business action, e.g., applying for licenses, initial regulatory setups, or entity formation. Note: There are almost no chances that One-time compliance will repeat.
                                        b. Ongoing Compliances: Requires continuous or recurring actions for the bank. Part of regular business operations, e.g., AML monitoring, KYC updates, regular reporting, Financial reporting, risk management, reason for rejection and compliance training.
                                        c. Event-based Compliance: Compliance actions that are triggered by specific events, e.g. In case of rejection of any Consumers' application for a Financial Product and/or Service by the Licensed Financial Institution, the institution must disclose the reason for rejection to the applicant, except where the reason for rejection is related to Financial Crime Compliance risks or as may be prohibited by law.
                                        5. Event Name: If the text mentions a bank's obligation and the compliance type is 'Event-based,' identify the specific event described in the text as Event Name. 
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                    Event Name: Reporting to RBI of an Account red flagged on the CRILC platform reporting threshold by the reporting entity.
                                        6. Event Applicability: If the text mentions a bank's obligation and the compliance type is 'Event-based,' describe the specific circumstances or conditions under which a compliance obligation must be fulfilled. It often ties the obligation to a particular event or occurrence.
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                    Event Applicability: Within 7 days of account being red flagged
                                        7. Event - Due Date: If the text mentions a bank's obligation and the compliance type is 'Event-based,' identify any due date, time or description available in the text as Event – Due Date. If you can't find anything, write 'X'.
                                            Example:
                                                    Compliance description/ Obligation: Report to the Reserve Bank within seven days of being red flagged an account meeting the CRILC reporting threshold by the reporting entity.
                                                Even Due Date: X+7 days
                                    {excel_column_data}"""
    
    
    return system_message


def create_system_message1(excel_column_data):
    system_message = f"""
       These instructions are for internal use only. Do not include them in the user-facing response.
                                    1. Ensure each action is instructive, beginning with verbs like ensure, maintain, adhere, confirm, establish, furnish, etc., in an active and assertive voice. If there are multiple actions, make sure each one begins with a verb
                                    2. Avoid terms such as IBU, IBUs, Bank, AD-Bank, Banking Company, Shall, May, etc., as much as possible.
                                    3. Make sure that you do not miss out on any information in the input text.
                                    4. Include the classification of the action at the end of the output in parentheses.
                                    {excel_column_data}"""
    
    
    return system_message

def get_obligation_api(paragraph, semaphore):
    #print("INSIDE 2 FOLD")
    
    #async with semaphore:
    if paragraph not in ("No actions.", "No actions", "No action","No action."):
        # delay_time = 10
        # time.sleep(delay_time)
        llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-flash-lite",          # Latest stable Gemini Flash – very fast & capable          # Free tier, fast & capable
        temperature=0.0,                    # Deterministic output for obligations
        max_output_tokens=2048,             # Enough for long responses
        google_api_key=os.environ["GOOGLE_API_KEY"]
        )


        prompt_template = f"""
        {{system_message}}
        Now, answer the following question based on the information provided:
        Question: {{question}}
        """

        prompt = PromptTemplate(template=prompt_template, input_variables=["system_message", "question"])

        responses = []
    
        
        system_message = create_system_message1(paragraph)
        final_prompt = prompt.format(
        system_message=system_message, 
        question = 'Each point should be instructive and it should start with verbs like "Ensure", "maintain", "adhere", "confirm", "establish", "furnish" etc. Please make sure that you do not miss any information. Include the classification of the action at the end of the output in parentheses. If "No actions." or "No action" is found, simply return "No actions."'
        #question="Each point should be instructive and it should start with verbs like \"Ensure\", \"maintain\", \"adhere\", \"confirm\", \"establish\", \"furnish\" etc. Please make sure that you do not miss any information. Include the classification of the action at the end of the output in parentheses. If \"No actions.\" or \"No action\" is found, simply return \"No actions.\""
        )
            
        messages = [
            {"role": "system", "content": system_message},
            {"role": "user", "content": final_prompt}
        ]
            
        response = llm.invoke(messages)
        responses.append(response.content)
        #print("RESPONSEE OF 2ND FOLD   +++++++++++",response)
    
    else:
        return "No Action."

    return responses


async def get_obligation_apiii(paragraph :str,access_token: str,semaphore) -> str:
    async with semaphore:
        if paragraph not in ("No actions.", "No actions", "No action","No action."):
            data=[]
            #url = "YOUR_AZURE_OPENAI_WEBSOCKET_URL"
            url = "YOUR_AZURE_OPENAI_WEBSOCKET_URL"
            my_dict = {}
            async with websockets.connect(url,timeout=50) as ws:
                out_message_dict={
                    
                            "ce-type": "user_message",
                            "query": [ 
                                {
                                    "role": "system",
                                    "content": """
                                    These instructions are for internal use only. Do not include them in the user-facing response.
                                    1. Ensure each action is instructive, beginning with verbs like ensure, maintain, adhere, confirm, establish, furnish, etc., in an active and assertive voice. If there are multiple actions, make sure each one begins with a verb
                                    2. Avoid terms such as IBU, IBUs, Bank, AD-Bank, Banking Company, Shall, May, etc., as much as possible.
                                    3. Make sure that you do not miss out on any information in the input text.
                                    4. Include the classification of the action at the end of the output in parentheses.
                                    """ + str(paragraph)
                                },
                                {
                                    "role": "user",
                                    "content": """
                                    Each point should be instructive and it should start with verbs like “Ensure”, "maintain", "adhere", "confirm", establish", furnish" etc. Please make sure that you do not miss any information.Include the classification of the action at the end of the output in parentheses. If "No actions." or "No action" is found, simply return "No actions."
                                    
                                    """
                                }
                            ],
                            "access_token": access_token
                            }
                out_message = json.dumps(out_message_dict)
                await ws.send(out_message)

                data = ""
                while True:
                    recv_message = await asyncio.wait_for(ws.recv(), timeout=20)
                    # Parse the received JSON message and extract the data
                    message_json = json.loads(recv_message)

                    if message_json.get("data", "") == "verbose end" or message_json.get("data", "") == "stream end":
                        break
                    data += message_json.get("data", "")

                return data.replace('"', '')
        else:
            'No actions.'

async def call_gpt1(list_of_paragraph):

    semaphore = asyncio.Semaphore(5)  # Limit to 5 concurrent connections
    #url, access_token = neg_api() 
    """ need to implement the logic for no action or we can check if it is having no action then update to no action"""
   
    tasks = [get_obligation_api1(paragraph,semaphore) for paragraph in list_of_paragraph]
    results = await asyncio.gather(*tasks)  # if async
    return tasks

async def call_gpt(list_of_paragraph):

    semaphore = asyncio.Semaphore(5)  # Limit to 5 concurrent connections
    #url, access_token = neg_api() 
    """ need to implement the logic for no action or we can check if it is having no action then update to no action"""
    tasks = [get_obligation_api(paragraph,semaphore) for paragraph in list_of_paragraph]
    results = await asyncio.gather(*tasks)  # if async
    return tasks


Regulator = "Reserve Bank of India (RBI)"

def string_match(myString1, myString2):
    return myString2 in myString1


def neg_api():
   
    azure_endpoint='YOUR_AZURE_NEGOTIATE_API_URL',
   
    headers = {
        "openai_api_version" :"YOUR_API_VERSION", 
        "openai_api_key" :'YOUR_API_KEY',
        "openai_api_type" :'YOUR_API_TYPE',
    }   
    payload={
    "temperature": "0.05",
    "top_p": "0.95",
    "max_tokens": "500",
    "presence_penalty": "0",
    "frequency_penalty": "0",
    "model_name": "YOUR_MODEL_NAME",
    }

    response = requests.post(azure_endpoint, json=payload, headers=headers)
    response.raise_for_status() 

    if response.status_code == 200:
        response_json=response.json()
        return response_json['baseUrl'],response_json['token']
        
    else:
        print("response_jeson: ", str(response.json()))
        print("Request to negotiate api failed")
        sys.exit(1)

def create_system_message(excel_column_data):
    system_message = f"""
      1. You are a chatbot who is an expert in Risk and Compliance in Banking domain. Your primary goal is to help compliance expert in a bank with information about the given data.
                                        2. You need to identify actions that {Applicability} are legally required to perform given the directions from various regulators like the {Regulator}.
                                        3. Please list down all these actions in detail that {Applicability} must carry out according to the documents given.
                                        4. Include only specific actions but not the generic ones. For e.g. the fact that {Applicability} are required to comply with the regulations in the RBI circular or guidance note is a generic action and is already known to a Compliance expert. You should highlight more specific actionable directly relating to each part of the regulation.
                                        5. Please ensure that a specific action that you have enlisted already is not mentioned again in a separate point.
                                        6. If these actions are dependent on actions by non-banking entities or individuals, then mention that context or dependency while you state the obligations.
                                        7. Please ensure that you cover all the actions mentioned in the specified document and list them down correctly without repetition.
                                        8. List down the actions only as per the uploaded documents in clear bullet points.
                                        9. If you find actions to be carried out by {Applicability}, strictly classify them only as Filing/Return or Display or Procedural Requirement. 
                                        “Filing/Return” in the context of banking and financial institutions typically refers to the submission of certain required documents or reports to regulatory authorities like the {Regulator}.An action will be classified as Filing/Return if {Applicability} is required to notify or submit or furnish specific documents, forms, facts, returns, reports statements, information back to the {Regulator}. Please provide the frequency only in case of Filing/Return.
                                        An obligation will be classified as Display if {Applicability} need to publicly disclose information.
                                        An obligation will be classified as Procedural Requirement if it is neither Filing/Return or Display.
                                        10. Please list only the actions to be carried out by {Applicability}. If there are no such actions, state that "No actions.".
                                        11. Always assign the classification with highest match score only, resulting one classification per action item.
                                        12. The classification of action should always be mentioned at the end of the output and in parentheses.
                                        13. Strictly stick to the text provided to you while generating obligations. Do not take information from other columns to create obligations. For example do not explicitly mention to whom the circular is applicable to while generating obligations.
                                        {excel_column_data}"""
    
    
    return system_message


from tqdm import tqdm

async def hello_a(list_of_paragraph: List[str], semaphore, applicability: str = "nan") -> List[str]:
    """
    Process list of paragraphs asynchronously using Gemini 2.5 Flash.
    Returns list of responses.
    """
    # ── Build LLM using current key in pool ────────────────────────────────
    global _current_key_index

    def _make_llm(api_key: str):
        return ChatGoogleGenerativeAI(
            model="gemini-2.5-flash-lite",
            temperature=0.0,
            max_output_tokens=2048,
            google_api_key=api_key,
        )

    llm = _make_llm(GOOGLE_API_KEYS[_current_key_index])
    print(f"  [Key pool] Starting with key index {_current_key_index}")

    applicability = applicability if applicability != 'nan' else "UAE Information Assurance Regulation"

    # ── System prompt: role + strict rules ──────────────────────────────────
    system_prompt = (
        f"You are a regulatory compliance expert analysing the UAE Information Assurance "
        f"Regulation and similar UAE government regulatory documents. "
        f"Your task is to extract every compliance obligation that {applicability} must perform. "
        f"STRICT RULE: Only extract sentences that contain 'shall', 'must', 'is required to', "
        f"'are required to', 'obliged to', or 'needs to' as the core mandatory verb. "
        f"NEVER extract sentences where 'should', 'may', or 'can' is the only modal verb — "
        f"those are guidance, not obligations. "
        f"NEVER invent, paraphrase, or generalise. Use the document's exact words only."
    )

    # ── Extraction instruction (user turn) ───────────────────────────────────
    extraction_instruction = """Extract ALL compliance obligations from the regulatory text below.

DOCUMENT STRUCTURE:
This UAE regulatory PDF uses a two-row table structure for every control:
  Row labelled 'Control':     A brief high-level "shall" statement.
  Row labelled 'Sub-Control': A detailed "The entity shall: 1) ... 2) ... 3) ..." list.

RULES — read carefully:

1. MANDATORY LANGUAGE ONLY:
   Only extract sentences whose main verb is 'shall', 'must', 'is required to',
   'are required to', 'obliged to', or 'needs to'.
   NEVER extract sentences where 'should', 'may', or 'can' is the only modal verb.
   'Should' sentences are guidance — skip them entirely.

2. CONTROL ROW = SEPARATE OBLIGATION:
   Every 'Control | The entity shall [action].' sentence is its own bullet.
   Every 'Sub-Control | The entity shall: 1)... 2)...' list is its own bullet.
   NEVER merge Control + Sub-Control into one bullet.
   NEVER drop the Control sentence because a Sub-Control follows it.

3. COMPLETE SENTENCES ONLY:
   A bullet must be a grammatically complete obligation — subject + shall/must + object.
   If a sentence appears cut off (ends with 'and', 'for', 'the', or a dangling verb),
   do NOT include it as a standalone bullet. Instead incorporate it with the next
   related sentence if they clearly belong together, or skip if unresolvable.
   NEVER output a bullet that is just 'The entity shall:' with nothing after the colon.

4. SUB-POINTS MUST STAY WITH THEIR PARENT:
   When a 'shall' statement is followed by numbered sub-points (1), 2), 3)... or a), b)...)
   that explain HOW to fulfil it, keep them together in ONE bullet.
   Write the main 'shall' statement, then incorporate the numbered sub-points naturally.
   Example: "The entity shall establish an asset management policy. The policy shall:
   be appropriate to the entity's purpose, outline roles and responsibilities, and
   be reviewed at planned intervals." (Procedural Requirement)

5. EXACT WORDING:
   Use the document's exact words. If the grammatical subject is missing (sentence starts
   with 'shall'), add it in square brackets: '[The entity] shall establish...'

6. CLASSIFICATION — end every bullet with one of these in parentheses:
   (Procedural Requirement) — mandatory action or implementation
   (Filing/Return)          — must submit, notify, or report to a regulator
   (Display)                — must publicly post or disclose information

7. If after careful reading there are genuinely NO 'shall/must' obligations,
   respond with exactly: No obligations.

8. No preamble, headers, or commentary. Bullet list only, one bullet per line starting with *.

REGULATORY TEXT:
{text}"""

    responses = []

    async with semaphore:
        for text in tqdm(list_of_paragraph, desc="Processing paragraphs"):
            try:
                user_message = extraction_instruction.format(text=text)
                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content": user_message},
                ]
                # ── Key-rotating invoke ────────────────────────────────
                keys_tried = 0
                success    = False
                while keys_tried < len(GOOGLE_API_KEYS):
                    try:
                        response = await llm.ainvoke(messages)
                        responses.append(response.content.strip())
                        success = True
                        break
                    except Exception as api_err:
                        err_str  = str(api_err)
                        is_quota = ("429" in err_str
                                    or "quota" in err_str.lower()
                                    or "ResourceExhausted" in err_str)
                        if is_quota and len(GOOGLE_API_KEYS) > 1:
                            keys_tried += 1
                            next_key = _get_next_key()
                            print(f"  [Key pool] 429 hit → rotating to key index {_current_key_index}")
                            llm = _make_llm(next_key)
                            await asyncio.sleep(2)
                        elif is_quota:
                            # Only one key — wait for quota window to reset
                            print(f"  [Key pool] Quota hit, waiting 65s for reset...")
                            await asyncio.sleep(65)
                            keys_tried = 0   # retry same key after wait
                            break
                        else:
                            print(f"  [Key pool] Non-quota error: {err_str[:120]}")
                            break

                if not success:
                    responses.append("Error: Could not process this paragraph.")

            except Exception as e:
                print(f"  Unexpected error on paragraph: {str(e)[:120]}")
                responses.append("Error: Could not process this paragraph.")

            await asyncio.sleep(2)

    return responses



async def hello_aaa(paragraph, access_token, semaphore):
    async with semaphore:
        data=[]
   

        url = "YOUR_AZURE_OPENAI_WEBSOCKET_URL"
        #url = "YOUR_AZURE_OPENAI_WEBSOCKET_URL"
        my_dict = {}
        # try:
        async with websockets.connect(url, timeout=100) as ws:  # Increase timeout to 100 seconds

            out_message_dict={
                
                        "ce-type": "user_message",
                        "query": [ 
                            {
                                "role": "system",
                                "content": f"""
                                        1. You are a chatbot who is an expert in Risk and Compliance in Banking domain. Your primary goal is to help compliance expert in a bank with information about the given data.
                                        2. You need to identify actions that {Applicability} are legally required to perform given the directions from various regulators like the {Regulator}.
                                        3. Please list down all these actions in detail that {Applicability} must carry out according to the documents given.
                                        4. Include only specific actions but not the generic ones. For e.g. the fact that {Applicability} are required to comply with the regulations in the RBI circular or guidance note is a generic action and is already known to a Compliance expert. You should highlight more specific actionable directly relating to each part of the regulation.
                                        5. Please ensure that a specific action that you have enlisted already is not mentioned again in a separate point.
                                        6. If these actions are dependent on actions by non-banking entities or individuals, then mention that context or dependency while you state the obligations.
                                        7. Please ensure that you cover all the actions mentioned in the specified document and list them down correctly without repetition.
                                        8. List down the actions only as per the uploaded documents in clear bullet points.
                                        9. If you find actions to be carried out by {Applicability}, strictly classify them only as Filing/Return or Display or Procedural Requirement. 
                                        “Filing/Return” in the context of banking and financial institutions typically refers to the submission of certain required documents or reports to regulatory authorities like the {Regulator}.An action will be classified as Filing/Return if {Applicability} is required to notify or submit or furnish specific documents, forms, facts, returns, reports statements, information back to the {Regulator}. Please provide the frequency only in case of Filing/Return.
                                        An obligation will be classified as Display if {Applicability} need to publicly disclose information.
                                        An obligation will be classified as Procedural Requirement if it is neither Filing/Return or Display.
                                        10. Please list only the actions to be carried out by {Applicability}. If there are no such actions, state that "No actions.".
                                        11. Always assign the classification with highest match score only, resulting one classification per action item.
                                        12. The classification of action should always be mentioned at the end of the output and in parentheses.
                                        13. Strictly stick to the text provided to you while generating obligations. Do not take information from other columns to create obligations. For example do not explicitly mention to whom the circular is applicable to while generating obligations.
                                        input data:""" + str(paragraph)
                            },
                            {
                                "role": "user",
                                # "content": f"""Could you please extract and list all the actions if any, required for {Applicability} from the uploaded regulatory document, using the exact same language as in the document? In case there are actions then also strictly classify them only as Filing/Return or Display or Procedural Requirement. If there are no such actions, state that, "No actions". Please provide the frequency in case of Filing/Return. Do not include {Applicability} in the obligations."""
                                "content": f"""Could you please extract and list all the actions if any, required for {Applicability} from the uploaded regulatory document, using the exact same language as in the document? In case there are actions then also classify them as Filing/Return or Display or Procedural Requirements. If there are no such actions, state that, "No actions.". Please provide the frequency in case of Filing/Return. Only list the obligations without generating any extra text after listing the obligations. Do not explicitly mention to whom the circular is applicable to while generating obligations, For Example - "This is a requirement for all Chairpersons and CEOs of Regulated Entities"."""
                            }
                        ],
                        "access_token": access_token
                        }
            #prompt = PromptTemplate(template=prompt_template, input_variables=["system_message", "question"])
            out_message = json.dumps(out_message_dict)
            await ws.send(out_message)

            data = ""
            while True:
                # recv_message = await asyncio.wait_for(ws.recv(), timeout=20)
                recv_message = await asyncio.wait_for(ws.recv(), timeout=30)  # Increase to 30 seconds

                # Parse the received JSON message and extract the data
                message_json = json.loads(recv_message)

                if message_json.get("data", "") == "verbose end" or message_json.get("data", "") == "stream end":
                    break
                data += message_json.get("data", "")

            return data.replace('"', '')
   

 
async def test_one(list_of_paragraph):
    semaphore = asyncio.Semaphore(5)  # Limit to 5 concurrent connections
    
    # Fixed: Added await + asyncio.gather since hello_a is async and returns list
    result = await hello_a(list_of_paragraph, semaphore)
    
    return result

import asyncio
import aiohttp

async def test_one_new(list_of_paragraph):
    semaphore = asyncio.Semaphore(5)  # Limit to 5 concurrent connections
    url, access_token = neg_api()

    async def fetch_with_retry(paragraph):
        while True:
            try:
                return await hello_a(paragraph, access_token, semaphore)  # ← Added await
            except aiohttp.ClientTimeout:
                print("Timeout error, regenerating access token...")
                url, access_token = neg_api()  # Regenerate token on timeout

    try:
        tasks = [asyncio.create_task(fetch_with_retry(paragraph)) for paragraph in list_of_paragraph]
        results = await asyncio.gather(*tasks)
    except Exception as e:
        print(f"An error occurred: {e}")
    return results

def classify_obligation(obligation):
    try:
        # Convert to lowercase and clean the text
        obligation = obligation.lower()

        # Regex to find the classification
        classification_pattern = r'classification:\s*(filing/return|display|procedural requirement)'
        classifications = re.findall(classification_pattern, obligation)

        # Capitalize the classifications found by regex
        if classifications:
            return ', '.join(set(classification.capitalize() for classification in classifications))

        # Check for keywords in the raw text
        keywords = ['procedural requirement', 'filing/return', 'display']
        found_keywords = [keyword.capitalize() for keyword in keywords if keyword in obligation]

        if found_keywords:
            return ', '.join(found_keywords)

        return 'N/A'

    except Exception as e:
        return f"Error occurred: {str(e)}"
    

def check_obligations(probable_obligations, comparison_variable):
    
    cleaned_probable_obligations = probable_obligations.strip().lower().replace(".", "").replace(" ", "")
    cleaned_comparison_variable = comparison_variable.strip().lower().replace(".", "").replace(" ", "")

    ret_string = string_match(cleaned_probable_obligations,cleaned_comparison_variable)
    if ret_string == True:
        result = "No"
    else:
        result = "Yes"
    return result




from typing import List

def update_no_action(obligation):
    apology_phrases = ['sorry', 'apologise', 'apologize', 'apologies', 'chatbot', 'no actions', 'no action', 'no actions.']
    
    # Convert obligation to lowercase for case-insensitive matching
    obligation_lower = obligation.lower()

    # Check if any apology phrase is in the obligation
    if any(phrase in obligation_lower for phrase in apology_phrases):
        return "No actions.", True
    
    return obligation, False





from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
def get_tab_obl(get_pagenumber, data):

    if pd.notna(get_pagenumber):
        page_reference = extract_page_reference(get_pagenumber)

        if page_reference:
            match = re.match(r'Page_(\d+)(_(\d+))?', page_reference)
            if match:
                page_number = int(match.group(1))
                item_index = int(match.group(3)) - 1 if match.group(3) else 0  # Default to 0 if no index
                
                # Keep trying to find a match for the page number
                while True:
                    found = False
                    for page_data in data:
                        if str(page_number) in page_data:
                            page_content = page_data[str(page_number)]
                            if 0 <= item_index < len(page_content):
                                table_val = page_content[item_index][1]
                                print(f"Extracted item from page {page_number}: {table_val}")
                                return table_val
                            else:
                                print(f"Item index {item_index} out of range for page {page_number}.")
                                return 'NA'
                            found = True
                            break
                    if not found:
                        print(f"Page number {page_number} not found. Trying next page.")
                        page_number += 1
                        if page_number > len(data):  # Stop if page_number exceeds available data
                            print(f"Exceeded maximum page number. No match found.")
                            return 'NA'
                    
            else:
                print(f"No match found for page reference {page_reference}.")
                return 'NA'
        else:
            print("Page reference could not be extracted.")
            return 'NA'
    else:
        print("Page number is not valid.")
        return 'NA'


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> External Function Start >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def process_obligations(df):

    def split_obligation(text):
        if isinstance(text, str):
           # pattern = r'(?<!\d)(^\d{1,2}\.\s+)'
           # pattern = r'(?<!\d)(\d{1,2}\.\s+)'
            pattern = r'(?:(?<=\n)|(?<=^))(\d{1,2}\.\s+|-\s+)'
            parts = re.split(pattern, text)
            if parts[0].strip() == '':
                parts = parts[1:]
            
            obligations = [parts[i] + parts[i+1] for i in range(0, len(parts) - 1, 2)]

            if len(parts) % 2 == 1:
                obligations.append(parts[-1].strip())
        else:
            obligations = [text]
        
        return obligations

    def expand_obligation(df):
        rows = []
        for _, row in df.iterrows():
            obligations = split_obligation(row['Expected Obligations'])
            for obligation in obligations:
                new_row = row.copy()
                new_row['Obligations Breakdown'] = obligation
                rows.append(new_row)
        return pd.DataFrame(rows)
    
    df_expanded = expand_obligation(df)
    print(f"Counter Break Process completed")
    return df_expanded

def process_gap_breakdown(df):

    def process_text(text):
        lines = text.split('\n')
        processed_lines = []
        buffer = []

        for line in lines:
            trimmed_line = line.strip()
            is_top_level_header = re.match(r'^\d+\.\s', trimmed_line)
            is_pointer_or_sub_header = re.match(
                r'^\d+(\.\d+)*\s|^\([ivxlc]+\)|^[a-z]\)|^[A-Z]\)|^\([a-z]\)|^\([A-Z]\)|^[ivxlc]+\.\s|^[a-z]\.\s|^[A-Z]\.\s|^[i]\)\s|^[ii]\)\s|^[iii]\)\s|^[iv]\)\s|^[v]\)\s|^[vi]\)\s|^[vii]\)\s|^[viii]\)\s|^[ix]\)\s|^[a]\)\s|^[b]\)\s|^[c]\)\s|^[d]\)\s|^[e]\)\s|^[f]\)\s|^[g]\)\s|^[h]\)\s',
                trimmed_line
            )

            if is_top_level_header:
                if buffer:
                    processed_lines.append(' '.join(buffer))
                    buffer = []
                processed_lines.append(trimmed_line)
                processed_lines.append('')
            elif is_pointer_or_sub_header:
                if buffer:
                    processed_lines.append(' '.join(buffer))
                    buffer = []
                buffer.append(trimmed_line)
            else:
                buffer.append(trimmed_line)

        if buffer:
            processed_lines.append(' '.join(buffer))

        result_text = '\n'.join(processed_lines)

        final_lines = []
        lines = result_text.split('\n')
        for line in lines:
            if re.match(r'^\d+\.\s', line):
                final_lines.append(line)
            else:
                if final_lines and final_lines[-1] == '':
                    final_lines[-1] = line
                else:
                    final_lines.append(line)

        return '\n'.join(final_lines)
    df['Regulatory Text BreakDown'] = df['Regulatory Text BreakDown'].apply(process_text)
    print("GAP Breakdown Process completed")
    return df 

def OT_Breakdown(df):
# Define the types to check (with proper capitalization)
    obligation_types = ['Procedural Requirement', 'Filing/Return', 'Display']

    def extract_obligation_type(text):
        """Extract all obligation types enclosed in parentheses from the text."""
        if pd.isna(text):
            return None
        # Find all types enclosed in parentheses
        matches = re.findall(r'\(([^)]+)\)', text)
        if matches:
            # Return all found types as a list
            return [match.strip() for match in matches]
        return []

    def normalize_type(extracted_type):
        """Normalize extracted type to match defined types."""
        # Convert to lower case and strip unwanted characters
        normalized_type = re.sub(r'\s*[-:]\s*', ' ', extracted_type).strip().lower()
        return normalized_type

    def find_matching_obligations(extracted_types, type_list):
        """Find and return the unique matching obligations from the extracted types."""
        # Convert type list to lower case for case-insensitive comparison
        type_list_lower = [ob_type.lower() for ob_type in type_list]
        
        # Find matches and ensure uniqueness
        matched_types = set()
        for extracted_type in extracted_types:
            # Normalize the extracted type
            normalized_type = normalize_type(extracted_type)
            for type_option in type_list_lower:
                if normalized_type.startswith(type_option.lower()):
                    matched_types.add(type_list[type_list_lower.index(type_option)])
        
        return ', '.join(sorted(matched_types)) if matched_types else None

    # Process each row in the DataFrame
    results = []
    for index, row in df.iterrows():
        # Handle missing values
        breakdown_text = row.get('Obligations Breakdown', '')
        
        # Extract all obligation types from breakdown text
        extracted_types = extract_obligation_type(breakdown_text)
        
        # Find matching obligations
        matching_obligations = find_matching_obligations(extracted_types, obligation_types)
        
        results.append(matching_obligations)

    # Add results to the DataFrame
    df['Obligations Type'] = results
    df['Obligations Type'].fillna("N/A", inplace=True)
    print(f"Obligation Type Breakdown Processing complete.")

    return df

def add_frequency_and_due_date(df):
    frequency_mapping = {
        "Monthly":"Banks must comply with the circular's instructions by the 10th of each month. Penalties will apply for late compliance, regardless of the issue date.",
        #"Monthly": "10th of the following month which it relates to",
        "Quarterly": "31 January, 30 April, 31 July, 31 October",
        "Semi-annually": "31 July, 31 January",
        "Annually": "31 January"
    }


    
    frequency_patterns = {
        key: re.compile(rf'\b{key}\b', re.IGNORECASE) for key in frequency_mapping.keys()
    }
    
    def tag_frequency_and_due_date(row):
        # Check for "No actions." in "Obligations Type"
        if "N/A" in row['Obligations Type']:
            return "NA", "NA"
        
        # Otherwise, process "Obligations Breakdown"
        obligation_text = row['Obligations Breakdown']
        
        # Check for frequency patterns
        for key, pattern in frequency_patterns.items():
            if pattern.search(obligation_text):
                return key, frequency_mapping[key]
        
        # Default to "Monthly" if no match is found
        return "Monthly", frequency_mapping["Monthly"]
    
    df['Frequency'], df['Due date'] = zip(*df.apply(tag_frequency_and_due_date, axis=1))
    print(f"Frequency Process completed")
    return df

def Text_removal(df):
    obligation_types = ['Procedural Requirement', 'Filing/Return', 'Display']

    def extract_obligation_type(text):
        if pd.isna(text):
            return []
        matches = re.findall(r'\(([^)]+)\)', text)
        return [match.strip() for match in matches] if matches else []

    def normalize_type(extracted_type):
        if extracted_type:
            normalized_type = re.sub(r'\s*[-:]\s*', ' ', extracted_type).strip().lower()
            return normalized_type
        return ""

    def find_matching_obligations(extracted_types, type_list):
        if not extracted_types:
            return None
        type_list_lower = [ob_type.lower() for ob_type in type_list]
        matched_types = set()
        for extracted_type in extracted_types:
            normalized_type = normalize_type(extracted_type)
            for type_option in type_list_lower:
                if type_option in normalized_type:
                    matched_types.add(type_list[type_list_lower.index(type_option)])
        return ', '.join(sorted(matched_types)) if matched_types else None

    def remove_trailing_text(text):
        if pd.isna(text):
            return text
        obligation_positions = [m.start() for m in re.finditer(r'\(([^)]+)\)', text)]
        if not obligation_positions:
            return text
        last_obligation_pos = max(obligation_positions)
        cleaned_text = text[:last_obligation_pos + text[last_obligation_pos:].find(')') + 1].strip()
        return cleaned_text

    def process_column(column_text):
        try:
            extracted_types = extract_obligation_type(column_text)
            matching_obligations = find_matching_obligations(extracted_types, obligation_types)
            cleaned_text = remove_trailing_text(column_text)
            return matching_obligations, cleaned_text
        except Exception as e:
            print(f"Error processing column text '{column_text}': {e}")
            return None, column_text  # Return None for obligations and original text

    # Initialize results lists
    results_obligations = []
    cleaned_texts_obligations = []
    results_expected_obligations = []
    cleaned_texts_expected_obligations = []
    results_additional_obligations = []
    cleaned_texts_additional_obligations = []

    # Process each row in the DataFrame
    for index, row in df.iterrows():
        try:
            breakdown_text = row.get('Obligations Breakdown', '')
            obligations_type, cleaned_breakdown = process_column(breakdown_text)
            results_obligations.append(obligations_type)
            cleaned_texts_obligations.append(cleaned_breakdown)
        except Exception as e:
            print(f"Error processing 'Obligations Breakdown' at row {index}: {e}")

        try:
            expected_obligations_text = row.get('Expected Obligations', '')
            expected_obligations_type, cleaned_expected_obligations = process_column(expected_obligations_text)
            results_expected_obligations.append(expected_obligations_type)
            cleaned_texts_expected_obligations.append(cleaned_expected_obligations)
        except Exception as e:
            print(f"Error processing 'Expected Obligations' at row {index}: {e}")

        try:
            obligations_text = row.get('Obligations', '')
            additional_obligations_type, cleaned_obligations = process_column(obligations_text)
            results_additional_obligations.append(additional_obligations_type)
            cleaned_texts_additional_obligations.append(cleaned_obligations)
        except Exception as e:
            print(f"Error processing 'Obligations' at row {index}: {e}")

    # Update the DataFrame
    df['Obligations Breakdown'] = cleaned_texts_obligations
    df['Expected Obligations'] = cleaned_texts_expected_obligations
    df['Obligations'] = cleaned_texts_additional_obligations

    print("Additional Text removal Processing complete")
    return df


def process_probable_obligations(df):
   
    exact_keywords = "Ensure, Maintain, Adhere, Confirm, Establish, Furnish, etc"
    similarity_threshold = 80 

    def replace_probable_obligations(row):
        probable_obligations = str(row['Expected Obligations']).strip()
        obligations = str(row['Obligations']).strip()
        
        if fuzz.token_set_ratio(probable_obligations.lower(), exact_keywords.lower()) >= similarity_threshold:
            if obligations and obligations.lower() != ["no actions","no action","no actions.","no action."] :
                return obligations
            elif obligations.lower() == ["no actions","no action","no actions.","no action."]:
                return obligations
        
        if probable_obligations.lower() == ["no actions","no action","no actions.","no action."] and obligations:
            return obligations

        return row['Expected Obligations']

    df['Expected Obligations'] = df.apply(replace_probable_obligations, axis=1)

    return df

def combine_no_actions_text(df):
    print("INSIDE COMBINE NO ACTION TEXT")
    combined_text = ""
    start_idx = None
    for i in range(len(df)):
        current_header = df.at[i, 'Section Chapter Header']
        #is_no_actions = df.at[i, 'Expected Obligations'] == "No actions." or "No action"
        is_no_actions = df.at[i, 'Expected Obligations'] == "No actions." or df.at[i, 'Expected Obligations'] == "No actions"

        is_new_header = (i == 0) or (current_header != df.at[i - 1, 'Section Chapter Header'])

        if is_new_header:
            if start_idx is not None and combined_text:
                #df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text.strip()
                df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text
            combined_text = ""
            start_idx = None

        if is_no_actions:
            if start_idx is None:
                start_idx = i  
            combined_text += " \n\n " + df.at[i, 'Regulatory Text BreakDown']
            df.at[i, 'Regulatory Text BreakDown'] = None
        else:
            if combined_text:
                #df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text.strip()
                df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text
                combined_text = ""
                start_idx = None
    if start_idx is not None and combined_text:
        #df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text.strip()
        df.at[start_idx, 'Regulatory Text BreakDown'] = combined_text
    df = df.dropna(subset=['Regulatory Text BreakDown'])
   # df.reset_index(drop=True, inplace=True)
    print("COMBINE NO ACTION TEXT COMPLETED")
    return df

def clean_text_removebefore1(text):
    if pd.isna(text):  # Check if the value is NaN
        return text  # Return NaN unchanged
    if "1." in text:
        split_text = text.split("1.", 1)  # Split the text at the first occurrence of "1."
        if split_text[0]:  # Check if there is text before "1."
            return "1." + split_text[1].strip()  # Return "1." and the text after it
        else:
            return text.strip()  # If nothing is before "1.", return original text without leading spaces
    return text  # Return the original text if "1." not found

def clean_dataframe(df):
    df['Expected Obligations'] = df['Expected Obligations'].apply(clean_text_removebefore1)
    return df
            
def replace_newline_digits_with_roman_in_df(df, columns):
    # Function to replace new lines followed by digits with small Roman numerals
    def replace_newline_digits_with_roman(text):
        if isinstance(text, str):
            roman_numerals = {
                '1.': 'i.',
                '2.': 'ii.',
                '3.': 'iii.',
                '4.': 'iv.',
                '5.': 'v.',
                '6.': 'vi.',
                '7.': 'vii.',
                '8.': 'viii.',
                '9.': 'ix.',
                '10.': 'x.',
                '11.': 'xi.',
                '12.': 'xii.',
                '13.': 'xiii.',
                '14.': 'xiv.',
                '15.': 'xv.',
                '16.': 'xvi.',
                '17.': 'xvii.',
                '18.': 'xviii.',
                '19.': 'xix.',
                '20.': 'xx.',

            }

            def replace_match(match):
                digit = match.group(0)
                #print("DIGIT:::",digit)
                return f'\n{roman_numerals[digit]}' if digit in roman_numerals else match.group(0)
            result = re.sub(r'^\d{1,3}\.', replace_match, text, flags=re.MULTILINE)
            return result.strip()
        else:
            return text
    # Apply the function to the specified columns
    for column in columns:
        df[column] = df[column].apply(replace_newline_digits_with_roman)
        df[column] = df[column].str.strip()
    print("Probable obligations and obligation breakdown changed from 1. to i.")
    return df

def remove_obligation_types(df):
    """
    Removes specified obligation types (e.g., (Procedural Requirement), (Filing/Return, Frequency: Post Facto), etc.)
    and any bracketed text at the end of sentences from the 'Expected Obligations' and 'Obligation Breakdown' columns.
    """
    print("REMOVE OBLIGATION TYPEEE")
    obligation_types = ['Procedural Requirement', 'Filing/Return', 'Display']
    def remove_obligation_type(text, types_list):
        if pd.isna(text):
            return text
        patterns = []
        for t in types_list:
            patterns.append(r'\(' + re.escape(t) + r'(?:,- [^)]+)?\)')
        general_pattern = r'\([^)]+\)'
        combined_pattern = '|'.join(patterns + [general_pattern])
        cleaned_text = re.sub(combined_pattern, '', text).strip()
        return cleaned_text
    df['Obligations Breakdown'] = df['Obligations Breakdown'].apply(lambda x: remove_obligation_type(x, obligation_types))
    df['Expected Obligations'] = df['Expected Obligations'].apply(lambda x: remove_obligation_type(x, obligation_types))
    df['Obligations Type'].fillna("N/A", inplace=True)
    df['Frequency'].fillna("NA", inplace=True)
    df['Due date'].fillna("NA", inplace=True)
    df['Log/Form Reference'].fillna("NA", inplace=True)
    print("Obligation type and bracket text removal processing complete.")
    return df


import pandas as pd

def replace_obligation(df):
    def update_obligation_type(row):
        # Check if 'Obligations Type' is missing or set to specific values and 'Obligations Breakdown' is not 'No actions.'
        is_obligation_type_missing = pd.isna(row['Obligations Type']) or row['Obligations Type'] in ['nan', 'N/A']
        is_no_action = row['Obligations Breakdown'] in ['No action.', 'No actions.','No action','No actions']
        
        # Return 'Procedural requirement' if conditions are met
        if is_obligation_type_missing and not is_no_action:
            return 'Procedural requirement'
        
        return row['Obligations Type']  # Return the original type if conditions are not met

    # Update 'Obligations Type' column based on the condition
    df['Obligations Type'] = df.apply(update_obligation_type, axis=1)
    
    # Replace NaN values with 'N/A' for 'Obligations Type' and 'NA' for all other columns
    df['Obligations Type'].fillna('N/A', inplace=True)
    df.fillna('NA', inplace=True)
    df.loc[df['Obligations Breakdown'].isin(['No action','No actions','No action.', 'No actions.']), 'Obligations Type'] = 'N/A'
    
    return df




def extract_compliance_details(df, output_column):
    # Check if the specified column exists
    if output_column not in df.columns:
        raise KeyError(f"The column '{output_column}' does not exist in the DataFrame.")
    
    # Initialize a dictionary to hold new column data
    new_columns = {
        'Compliance Classification': [],
        'Sub Classification': [],
        'Compliance Header': [],
        'Penalty Type': [],
        'Compliance Type (Timeframe)': [],
        'Event Name': [],
        'Event Applicability': [],
        'Event - Due Date': [],
    }
    
    def convert_to_string(x):
        if isinstance(x, list):  # If it's a list, join its elements into a single string
            return '\n'.join(x)
        return x 


    # Iterate through each row in the specified output column
    for text in df[output_column]:
        text = convert_to_string(text)
        if isinstance(text, str):  # Check if the entry is a string
            line_details = {
                'Compliance Classification': None,
                'Sub Classification': None,
                'Compliance Header': None,
                'Penalty Type': None,
                'Compliance Type (Timeframe)': None,
                'Event Name': None,
                'Event Applicability': None,
                'Event - Due Date': None,
            }
            
            lines = text.split('\n')
            for line in lines:
                if 'Compliance Classification:' in line:
                    line_details['Compliance Classification'] = line.split(': ')[1].strip().strip("'")
                if 'Sub Classification:' in line:
                    line_details['Sub Classification'] = line.split(': ')[1].strip().strip("'")
                if 'Compliance Header:' in line:
                    line_details['Compliance Header'] = line.split(': ')[1].strip().strip("'")
                if 'Penalty Type:' in line:
                    line_details['Penalty Type'] = line.split(': ')[1].strip().strip("'")
                if 'Compliance Type:' in line:
                    line_details['Compliance Type (Timeframe)'] = line.split(': ')[1].strip().strip("'")
                if 'Event Name:' in line:
                    line_details['Event Name'] = line.split(': ')[1].strip().strip("'")
                if 'Event Applicability:' in line:
                    line_details['Event Applicability'] = line.split(': ')[1].strip().strip("'")
                if 'Event Due Date:' in line:
                    line_details['Event - Due Date'] = line.split(': ')[1].strip().strip("'")

            # Append extracted details to the new_columns dictionary
            for key in new_columns:
                new_columns[key].append(line_details[key])
        else:
            # Append None for all new columns if the entry is not a string
            for key in new_columns:
                new_columns[key].append(None)

    # Add new columns to the DataFrame
    for key, value in new_columns.items():
        df[key] = value

    return df


def determine_values(obligation_type):
    if obligation_type in ['Procedural requirement', 'Procedural Requirement','Display']:
        return pd.Series(['Medium', 'No'])
    elif obligation_type == 'Filing/Return':
        return pd.Series(['High', 'Yes'])
    else:
        return pd.Series(['N/A', 'N/A'])







# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>.. External Function END >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



async def call_with_retry(chunk, retries=2):
    for attempt in range(retries):
        try:
            final_answers = await test_one(chunk)
            return final_answers  # Return if successful
        except Exception as e:
            print(f"An error occurred during API call first folder retrieve 1: {attempt} {e}")
            if attempt < retries - 1:
                await asyncio.sleep(0.5)  # Wait before retrying
                continue  # Retry
            else:
                print("Max retries exceeded for chunk. Dividing chunk and retrying.")
                # Divide the chunk into two parts
                mid_index = len(chunk) // 2
                first_half = chunk[:mid_index]
                second_half = chunk[mid_index:]
                # Call the API for the first half
                first_result = await call_with_retry(first_half, retries)
                # Call the API for the second half
                second_result = await call_with_retry(second_half, retries)
                print("DONE")
                return first_result + second_result

    # In case of failure before reaching the retries
    return ['NA'] * len(chunk)  # Fallback result


def call_with_retry(chunk, retries=2):
    for attempt in range(retries):
        try:
            final_answers = test_one(chunk)
            return final_answers  # Return if successful
        except Exception as e:
            print(f"An error occurred during API call first folder retrieve 1: {attempt} {e}")
            if attempt < retries - 1:
                #await asyncio.sleep(0.5)  # Wait before retrying
                continue  # Retry
            else:
                print("Max retries exceeded for chunk. Dividing chunk and retrying.")
                # Divide the chunk into two parts
                mid_index = len(chunk) // 2
                first_half = chunk[:mid_index]
                second_half = chunk[mid_index:]
                # Call the API for the first half
                first_result = call_with_retry(first_half, retries)
                # Call the API for the second half
                second_result = call_with_retry(second_half, retries)
                print("DONE")
                return first_result + second_result

    # In case of failure before reaching the retries
    return ['NA'] * len(chunk)  # Fallback result

async def process_data(data_dict, is_json, js_path):
    """
    Revised pipeline:
    - Reads table text (tab1.json) and joins split tabula cells into full sentences
    - Pre-filters rows with no mandatory language → instant 'No obligations.'
    - Only rows with real 'shall/must/required' go to LLM
    - SKIP sentinel used to preserve row order without extra LLM calls
    """
    paragraphs_to_process = []
    obligation_list       = []
    js_data               = []

    MANDATORY_RE  = re.compile(
        r'\b(shall|must|is required to|are required to|obliged to|will implement|needs to)\b', re.I
    )
    OBLIGATION_KW = ['shall', 'must', 'required', 'obliged', 'will implement', 'needs to']

    # ── helpers ──────────────────────────────────────────────────────────────

    def _join_table_text(raw_tab: str) -> str:
        """
        Tabula splits multi-line table cells into separate rows separated by
        newlines. Rejoin continuation lines (lines that do NOT start with a
        recognised label or capital-letter new-sentence) so that truncated
        'The entity shall ... [cut]\nand manage ...' becomes one full sentence.
        """
        labels = {'Control', 'Sub-Control', 'Implementation Guidance',
                  'Priority', 'Applicability', 'Unnamed'}
        lines  = raw_tab.splitlines()
        joined = []
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            first_word = stripped.split()[0].rstrip(':') if stripped.split() else ''
            is_label   = first_word in labels
            is_new_sent = stripped[0].isupper() and (
                stripped.startswith('The ') or stripped.startswith('Critical ')
                or stripped[0].isdigit() or stripped.startswith('[')
            )
            if joined and not is_label and not is_new_sent:
                joined[-1] = joined[-1].rstrip() + ' ' + stripped
            else:
                joined.append(stripped)
        return '\n'.join(joined)

    def _get_flat_table_text(page_int, next_section_page=None) -> str:
        """
        Return table text for a section, filtered at BLOCK level.

        A block = one Control row + its Sub-Control row + all numbered sub-points.
        Sub-points (1), 2), 3)...) do NOT contain 'shall' themselves but they are
        the actual content of what the entity shall do — they must be preserved.

        Filtering line-by-line would strip sub-points and give the LLM incomplete,
        context-free obligations. Block-level filtering keeps the whole block if
        ANY line in it contains mandatory language, so the LLM receives:
            "The entity shall establish a policy. The entity shall: 1)... 2)... 3)..."
        instead of just: "The entity shall establish a policy."

        Scan range: from page_int up to (not including) next_section_page.
        Works for any PDF structure — no section names or page offsets hard-coded.
        """
        if not js_data or page_int is None:
            return ''

        end_page = next_section_page if next_section_page else page_int + 8
        combined = []
        for pg in range(page_int, end_page):
            for table in js_data:
                if table.get('page') == pg:
                    cols      = table.get('columns', [])
                    rows      = table.get('rows',    [])
                    header    = ' | '.join(str(c) for c in cols)
                    row_lines = [' | '.join(str(cell) for cell in row) for row in rows]
                    combined.append(header + '\n' + '\n'.join(row_lines))

        if not combined:
            return ''

        raw    = '\n'.join(combined)
        joined = _join_table_text(raw)

        #  Block-level filter 
        BLOCK_STARTERS = ('Control', 'Sub-Control')
        blocks         = []
        current_block  = []

        for line in joined.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            first_word = stripped.split()[0].rstrip(':')
            if first_word in BLOCK_STARTERS and current_block:
                blocks.append(current_block)
                current_block = [stripped]
            else:
                current_block.append(stripped)
        if current_block:
            blocks.append(current_block)

        kept = []
        for block in blocks:
            block_text = ' '.join(block)
            if any(kw in block_text.lower() for kw in OBLIGATION_KW):
                kept.extend(block)

        return '\n'.join(kept).strip()

    # load JSON 
    if js_path and os.path.isfile(js_path):
        js_data = read_json(js_path) or []

    # build paragraphs 
    all_items = list(data_dict.values())
    page_ints = []
    for d in all_items:
        try:
            page_ints.append(int(float(str(d.get('sheet_name', '')))))
        except Exception:
            page_ints.append(None)

    for idx_d, data in enumerate(all_items):
        regulatory_text = data.get('regulatory_text', '') or ''
        page_number_raw = data.get('sheet_name', '')

        # Robustly parse page number
        try:
            page_int = int(float(str(page_number_raw)))
        except Exception:
            page_int = None

        # Next section's page = upper bound for table scan
        # This ensures we only read tables that belong to THIS section
        next_pg = None
        for future_pg in page_ints[idx_d + 1:]:
            if future_pg is not None and future_pg != page_int:
                next_pg = future_pg
                break

        # Table text: scan from this section's page up to next section's page
        filtered_table = _get_flat_table_text(page_int, next_pg) if js_data else ''

        # Pre-filter: CONTENT-ONLY check — no section names, no hard-coded rules
        combined_check = regulatory_text + ' ' + filtered_table
        if not MANDATORY_RE.search(combined_check):
            paragraphs_to_process.append('__SKIP__')
            continue

        # Build LLM input: reg text + table rows (clearly labelled)
        if filtered_table:
            raw_text = (
                regulatory_text.strip() + '\n\n'
                + '[CONTROL TABLE ROWS — extract obligations from these as well]:\n'
                + filtered_table
            )
        else:
            raw_text = regulatory_text.strip()

        paragraphs_to_process.append(raw_text)

    # ── send to LLM ───────────────────────────────────────────────────────────
    real_indices = [i for i, p in enumerate(paragraphs_to_process) if p != '__SKIP__']
    real_texts   = [paragraphs_to_process[i] for i in real_indices]
    llm_results  = []

    print(f"  Rows sent to LLM: {len(real_texts)} / {len(paragraphs_to_process)} total")

    chunk_size = 25
    for i in range(0, len(real_texts), chunk_size):
        chunk = real_texts[i:i + chunk_size]
        try:
            final_answers = await call_with_retry(chunk)
            llm_results.extend(
                [ans.replace('stream end', '') if isinstance(ans, str) else str(ans)
                 for ans in final_answers]
            )
            await asyncio.sleep(5)
        except Exception as e:
            print(f"Error during chunk processing: {e}")
            llm_results.extend(['NA'] * len(chunk))

    # Rebuild full list preserving row order
    result_iter = iter(llm_results)
    for para in paragraphs_to_process:
        if para == '__SKIP__':
            obligation_list.append('No obligations.')
        else:
            obligation_list.append(next(result_iter, 'NA'))

    return obligation_list


async def process_all_obligations(obligation_list, df):
    processed_obligations = []
    
    # Convert column to string to avoid FutureWarning
    df['Obligations'] = df['Obligations'].astype(str)
    
    for i, obligation in enumerate(obligation_list):
        # FIX: obligation is a LIST from Gemini (one per paragraph)
        # Convert to single string: join with newlines
        if isinstance(obligation, list):
            obligation_str = "\n".join(str(item).strip() for item in obligation if item)
        else:
            obligation_str = str(obligation).strip()
        
        # Now use the string version
        obligation_clean, has_no_action = update_no_action(obligation_str)
        obligation_type = classify_obligation(obligation_clean)  # ← safe, now string
        
        df.at[i, OBLIGATIONS_TYPE] = obligation_type
        processed_obligations.append((obligation_clean, obligation_type))
    
    df.fillna('NA', inplace=True)
    
    return processed_obligations


        
async def call_with_retry2(chunk, retries=2):
    for attempt in range(retries):
        try:
            final_answers = await call_gpt(chunk)
            return final_answers  # Return if successful
        except Exception as e:
            print(f"An error occurred during API call Second folder retrive 1:{attempt} {e}")
            if attempt < retries - 1:
                await asyncio.sleep(0.5)  # Wait before retrying
                continue  # Retry
            else:
                print("Max retries exceeded for chunk. Dividing chunk and retrying.")
                # Divide the chunk into two parts
                mid_index = len(chunk) // 2
                first_half = chunk[:mid_index]
                second_half = chunk[mid_index:]
                # Call the API for the first half
                first_result = await call_with_retry2(first_half, retries)
                # Call the API for the second half
                second_result = await call_with_retry2(second_half, retries)
                print("DONE")
                return first_result + second_result
    print("Max retries exceeded for chunk.")
    # In case of failure before reaching the retries
    #return ['NA'] * len(chunk)
    return ['No actions'] * len(chunk)



async def call_with_retry3(chunk, retries=2):
    for attempt in range(retries):
        try:
          
            final_answers1 = await call_gpt1(chunk)
            return final_answers1  # Return if successful
        except Exception as e:
            print(f"An error occurred during API call Second folder retrive 1:{attempt} {e}")
            if attempt < retries - 1:
                await asyncio.sleep(0.5)  # Wait before retrying
                continue  # Retry
            else:
                print("Max retries exceeded for chunk. Dividing chunk and retrying.")
                # Divide the chunk into two parts
                mid_index = len(chunk) // 2
                first_half = chunk[:mid_index]
                second_half = chunk[mid_index:]
                # Call the API for the first half
                first_result = await call_with_retry3(first_half, retries)
                # Call the API for the second half
                second_result = await call_with_retry3(second_half, retries)
                # print("DONE")
                return first_result + second_result
    print("Max retries exceeded for chunk.")
    # In case of failure before reaching the retries
    return ['NA'] * len(chunk)



async def update_obligations1(df, excel_file_path,obligation_list,file_name):
    print("INSIDE UPDATE OBLIGATION")
    
    #processed_obligations = await process_all_obligations(obligation_list, df)
    processed_obligations = obligation_list
    expected_obligation1 = []
    print("PROCESSED OBLIGATIONS COMPLETED")
   
    
    if processed_obligations:
        print("Inside Junk code")
        chunk_size = 15
        for i in range(0, len(processed_obligations), chunk_size):
            chunk = processed_obligations[i:i + chunk_size]
            try:
                #print("CHUNK-----",chunk)
               
                final_answers1 = await call_with_retry3(chunk)  # Use the retry function
                expected_obligation1.extend([ans for ans in final_answers1])
            except Exception as e:

                print(f"An error occurred during Second fold after retring: {e}")
                expected_obligation1.extend(['NA'] * len(chunk))
    else:
        expected_obligation1.extend(['NA'] * len(processed_obligations))
    expected_length = len(df)
    if len(expected_obligation1) < expected_length:
        expected_obligation1.extend(['N/A'] * (expected_length - len(expected_obligation1)))
    elif len(expected_obligation1) > expected_length:
        expected_obligation1 = expected_obligation1[:expected_length]
    try:
        df.insert(df.columns.get_loc(OBLIGATIONS_BREAKDOWN) + 1, OUTPUT, expected_obligation1)
        #df.to_excel("Doneee2.xlsx",index=False)
        print("Printing output column::::",df["OUTPUT"])
      
    except: 
        pass
    return df

async def update_obligations(df, excel_file_path,obligation_list,file_name):
    print("INSIDE UPDATE OBLIGATION")
    
    processed_obligations = await process_all_obligations(obligation_list, df)
    expected_obligation = []



    if processed_obligations:
        chunk_size = 15
        for i in range(0, len(processed_obligations), chunk_size):
            chunk = processed_obligations[i:i + chunk_size]
            
            try:
                final_answers = await call_with_retry2(chunk)  # Use the retry function

                #expected_obligation.extend([ans.replace(' ', '') if ans is not None else 'No actions' for ans in final_answers])
                expected_obligation.extend([ans.replace('stream end', '') if isinstance(ans, str) else 'No actions' for final_answer in final_answers for ans in final_answer])

            except Exception as e:

                print(f"An error occurred during Second fold after retring: {e}")
                expected_obligation.extend(['NA'] * len(chunk))
    else:
        expected_obligation.extend(['NA'] * len(processed_obligations))

    try:
        # df.insert(df.columns.get_loc(OBLIGATIONS), EXPECTED_OBLIGATIONS, expected_obligation)
        # Adding the new column "Expected Obligations" after "Obligations"
        df.insert(df.columns.get_loc(OBLIGATIONS) + 1, EXPECTED_OBLIGATIONS, expected_obligation)

    except: 
        pass
    prev_df =False
    ########################################
   
    ##################################################
    try:
    # Clean the DataFrame to remove any text that is present before 1. and handle missing values
        df['Expected Obligations'] = df['Expected Obligations'].apply(lambda x: 'No actions' if x == 'nan' else x)
        #df.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\first.xlsx")
        df10 = clean_dataframe(df)
        df10['Obligations Type'].fillna("N/A", inplace=True)
        df10['Frequency'].fillna("NA", inplace=True)
        df10['Log/Form Reference'].fillna("NA", inplace=True)
 
        try:
            df16 = combine_no_actions_text(df10)
            #df16.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\third.xlsx")
        except Exception as e:
            print(f"Failed to combine no actions text: {e}")
            df16 = df10  # Fallback to previous DataFrame
        try:
            df6 =  process_probable_obligations(df16)
            #df6.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\A.xlsx")
        except Exception as e:
            print(f"Failed to process obligations: {e}")
            df6 = df16  # Fallback to previous DataFrame

        # Obligations breakdown is taking place at digit. and - and a new column named Obligation breakdown is created 
        try:
            df1 = process_obligations(df6)
            #df1.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\B.xlsx")
        except Exception as e:
            print(f"Failed to process obligations: {e}")
            df1 = df6  # Fallback to previous DataFrame

        # Obligation type breakdown is done w.r.t Obligation breakdown column 
        try:
            df3 = OT_Breakdown(df1)
           # df3.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\C.xlsx")
        except Exception as e:
            print(f"Failed to breakdown obligations: {e}")
            df3 = df1  # Fallback to previous DataFrame

        # Add frequency and due date column 
        try:
            df4 = add_frequency_and_due_date(df3)
            #df4.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\D.xlsx")
        except Exception as e:
            print(f"Failed to add frequency and due date: {e}")
            df4 = df3  # Fallback to previous DataFrame

        # Text removal [Removing extra ()] Removing text in () which is present at end assuming it be obligation type
        try:
            df5 = Text_removal(df4)
           # df5.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\E.xlsx")
        except Exception as e:
            print(f"Failed to remove text: {e}")
            df5 = df4  # Fallback to previous DataFrame

        # Remove obligation types Removing Procedureal, Filing/Retrun, Display from Obligation text
        try:
            df7 = remove_obligation_types(df5)
           # df7.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\F.xlsx")
        except Exception as e:
            print(f"Failed to remove obligation types: {e}")
            df7 = df5  # Fallback to previous DataFrame

        # Replace obligation to procedural if some other obligation is found 
        try:
            df8 = replace_obligation(df7)
            # if df8:
            #     df8.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\G.xlsx")
        except Exception as e:
            print(f"Failed to replace obligation: {e}")
            df8 = df7  # Fallback to previous DataFrame
        # Replacing obligation that are 1.,2.,3. to i.,ii.,iii.
        try:
            df9 = replace_newline_digits_with_roman_in_df(df8,['Expected Obligations','Obligations Breakdown'])
            #df9.to_excel(r"C:\Users\JV238ZP\OneDrive - EY\Pictures\check\H.xlsx")
        except Exception as e:
            print(f"Failed to replace obligation: {e}")
            df9 = df8  # Fallback to previous DataFrame

        # Add Act ID
        df9["Act ID"] = range(1, len(df9) + 1)

        # Compliance ID handling
        terms_to_check = ['CHAPTER', 'SECTION', 'PART', 'chapter', 'section', 'part', 'MODULE']
        chapter_variation = df9['Compliance ID'].apply(lambda x: any(term in x for term in terms_to_check)).any()
        if not chapter_variation:
            df9['Compliance ID'] = ['RBI-' + str(i) for i in range(1, len(df9) + 1)]

        if "Circular Issue Date_CDSL" in  df9.columns:
                        df9.rename(columns={"Circular Issue Date_CDSL": "Circular Issue Date"}, inplace=True)
        if "Circular Number_CDSL" in  df9.columns:   
                        df9.rename(columns={"Circular Number_CDSL": "Circular Number"}, inplace=True)
        print("COLUMNS RENAMED")
  
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        df9 = df8  # Fallback to original DataFrame if any unexpected error occurs
    obligations_breakdown = df9["Obligations Breakdown"].to_list()
    #obligations_breakdown = df9["Obligations"].to_list()
    # df9.to_excel("AFTER DF9.xlsx",index= False)
    df10 = await update_obligations1(df9, excel_file_path, obligations_breakdown, file_name)
    # df10.to_excel("AFTER DF10.xlsx",index= False)
    df11 = extract_compliance_details(df10, OUTPUT)
    df11[['Criticality', 'Proof of Compliance']] = df11['Obligations Type'].apply(determine_values)
    df11['Regulatory Text BreakDown'] = df11['Regulatory Text BreakDown'].str.strip()
    # df11.to_excel("AFTER DF11.xlsx",index= False)
    try:
        
        wb = load_workbook(excel_file_path, data_only=False)
        sheet1 = wb['Sheet1']
        data_dict = {
            index: {
                'log_status': row.get(LOG_FORM_STATUS, ''),
                'sheet_name': row.get(PAGE_NUMBER, ''),
            }
            for index, row in df11.iterrows()
        }

        
        for index, obligation in enumerate(expected_obligation):
            if index in data_dict:
                log_status = data_dict[index]['log_status']
                sheet_name = data_dict[index]['sheet_name']

                if log_status == 'Yes':
                    log_text = f"Table found on page {sheet_name}"
                    hyperlink = f'{file_name}#{sheet_name}!A1'
                
                    df11.at[index, 'Log/Form Reference'] = f'=HYPERLINK("{hyperlink}", "{log_text}")'
            else:
                print("Key is not present in data dict")
    except Exception as e:
        print("Error in processing workbook {e}")

    # Fill NA values
    
    df11.fillna('NA', inplace=True)
    df11.loc[df9['Applicability'].isin(['N/A', 'NA']), 'Applicability'] = 'DEPOSITORY PARTICIPANTS (DPs)'
    try:
        df11['Name of Mother Act'] = df11['Mother Act']
        df11['Legislation / Act'] = df11['Mother Act']
    except:
        df11['Name of Mother Act'] = "NA"
        df11['Legislation / Act'] = "NA"
    df11['Statutory Authority'] = 'N/A'
    df11['Name of Rules'] = 'N/A'
    df11['Penalty Description '] = 'N/A'
    
    df11.rename(columns={'Due date': 'Due Date Information', 'Expected Obligations': 'Compliance Description',
                        'Act ID':'EY Act ID','Compliance ID':'EY Compliance ID',
                        'Circular Number':'Compliance/ Statutory Provisions',
                        'Regulatory Text BreakDown':'Additional Information/ Guidelines',
                        'Obligations Type':'Categorization'}, inplace=True)
    '''
    column_order_new = ['Act ID', 'Compliance ID', 'Central State Local', 'Mother Act', 
            'Circular Issue Date', 'Circular Number', 'Compliance Theme', 'Applicability', 
            'Section Chapter Header', 'Regulatory Text BreakDown',"Obligations", 'Probable Obligations','Obligations Breakdown' ,'Obligations Type', 
            'Frequency','Due Date Information','Log/Form Reference','Log/Form Status']
    '''
    

    column_order_new = ['EY Act ID', 'EY Compliance ID', 'Central State Local', 'Name of Mother Act', 
                        'Name of Rules','Legislation / Act','Circular Issue Date', 'Compliance/ Statutory Provisions','Compliance Theme', 
                        'Compliance Classification','Sub Classification', 'Categorization', 'Compliance Header','Applicability', 
                        'Section Chapter Header','Compliance Description', 'Additional Information/ Guidelines', 
                        'Obligations Breakdown', 'Log/Form Reference', 'Log/Form Status',
                        'Penalty Type','Penalty Description ','Compliance Type (Timeframe)','Frequency', 'Due Date Information',
                        'Statutory Authority','Criticality', 'Proof of Compliance','Event Name', 
                        'Event Applicability', 'Event - Due Date']
    
    # Check if there are duplicate columns
    if df11.columns.duplicated().any():
    # Remove duplicate columns
        df11 = df11.loc[:, ~df9.columns.duplicated()]
    df11[['Event Name', 'Event Applicability', 'Event - Due Date']] = df11[['Event Name', 'Event Applicability', 'Event - Due Date']].fillna('-').replace('NA', '-')
    df11['Additional Information/ Guidelines'] = df11['Additional Information/ Guidelines'].astype(str).str.strip()
    #df9 = df9.loc[:, ~df9.columns.duplicated()]
    df11 = df11.reindex(columns=column_order_new)
    df11.fillna('NA', inplace=True)
    print("COLUMN REINDEXING DONE")
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
        for cell in row:
            cell.value = None
    font = Font(name='Arial', size=10)

    # Write the updated DataFrame to Sheet1
    for r_idx, row in enumerate(dataframe_to_rows(df11, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            # if isinstance(value, list):
            #     value = " ".join(value)  # Join list items into a single string
            # elif isinstance(value, str) and '\n' in value:
            #     value = value.replace('\n', ' ')
            #sheet1.cell(row=r_idx, column=c_idx, value=value)
            cell = sheet1.cell(row=r_idx, column=c_idx, value=value)
            cell.font = font
    ########
    #####

    # Save the updated workbook
    wb.save(excel_file_path)
    print("Sheet2 updated successfully with expected obligation.")
    
    return df11
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> External END >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def _apply_obligations_styling(excel_file_path: str) -> None:
    """
    Re-apply professional styling after pandas df.to_excel() wipes openpyxl formatting.
    Navy header row, white bold text, alternating row colours, column widths, wrap text.
    Called automatically after every obligations save.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    NAVY        = "1F3864"   # header background
    LIGHT_BLUE  = "DCE6F1"   # alternate row tint
    WHITE       = "FFFFFF"
    BORDER_COL  = "B8CCE4"

    thin = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Column widths (A–I)
    col_widths = {1:28, 2:55, 3:10, 4:18, 5:38, 6:14, 7:12, 8:20, 9:80}

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, row in enumerate(ws.iter_rows(), start=1):
        is_header = (i == 1)
        is_odd    = (i % 2 == 1)

        for cell in row:
            # Fill
            if is_header:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            else:
                bg = LIGHT_BLUE if is_odd else WHITE
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size= 9, name="Calibri")

            # Alignment — wrap text for obligations column (col 9), others top-left
            if cell.column == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="left")
            cell.border = border

        # Row height: header taller, obligation rows auto
        if is_header:
            ws.row_dimensions[i].height = 22
        else:
            ws.row_dimensions[i].height = 80   # auto

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(excel_file_path)


def call_external_api(is_json, df, excel_file_path, file_name, js_path):

    # Prepare input for LLM
    data_dict = {
        index: {
            'regulatory_text': row.get(REGULATORY_TEXT_BREAKDOWN, ''),
            'log_status': row.get(LOG_FORM_STATUS, ''),
            'sheet_name': row.get('Page Number', ''),
        }
        for index, row in df.iterrows()
    }

    async def main():

        # Step 1: Get obligations from LLM
        obligations = await process_data(data_dict, is_json, js_path)

        print("Writing obligations into dataframe...")

        # Step 2: Ensure column is string type
        if 'Obligations' not in df.columns:
            df['Obligations'] = ""

        df['Obligations'] = df['Obligations'].astype(object)

        # Step 3: Write obligations (with bullet-level deduplication)
        def _dedup_bullets(text: str) -> str:
            """
            Remove duplicate bullet points from LLM output.
            Case-insensitive, strips trailing punctuation — catches near-identical
            duplicates that differ only in a comma vs semicolon.
            """
            if not isinstance(text, str) or not text.strip():
                return text
            seen      = []
            seen_norm = set()
            for line in text.split('\n'):
                norm = line.strip().lstrip('*\u2022-').strip().lower()
                norm = norm.rstrip('.,;')
                norm = ' '.join(norm.split())
                if norm and norm in seen_norm:
                    continue          # duplicate bullet — drop it
                seen_norm.add(norm)
                seen.append(line)
            return '\n'.join(seen)

        for index, obligation in enumerate(obligations):
            df.at[index, 'Obligations'] = _dedup_bullets(obligation)

        # Step 4: Fill NA safely
        df.fillna('NA', inplace=True)

        # ⭐ FINAL SAVE — write data then re-apply styling via openpyxl
        df.to_excel(excel_file_path, index=False)
        _apply_obligations_styling(excel_file_path)

        print("✅ Excel updated with obligations.")
        print("Styling applied to Excel output.")

        return obligations

    return asyncio.run(main())




import os
import pandas as pd
import time
import asyncio

def get_applicability_and_regulator(df):
    Applicability = df['Applicability'].iloc[0]
    if Applicability == "Not found":
        Applicability = "All banks"
    return Applicability

def rbi__ngo_api_2(excel_file_path, js_path=None, request_id=None):

    result = {}

    try:
        global Applicability

        file_name = os.path.basename(excel_file_path)

        # Load Excel
        df = pd.read_excel(excel_file_path)

        # JSON check
        if js_path:
            is_json = os.path.isfile(js_path)
        else:
            print("js_path is None. Proceeding without JSON file.")
            is_json = None

        # Applicability
        try:
            Applicability = get_applicability_and_regulator(df)
            print(f"Applicability :> {Applicability}")
        except:
            pass

        start = time.time()

        # ⭐ CALL LLM
        obligations = call_external_api(
            is_json,
            df,
            excel_file_path,  # ← IMPORTANT
            file_name,
            js_path
        )

        end = time.time()

        print("TIME FOR OBLIGATION EXTRACTION============================")
        print(f"Runtime: {end - start:.2f} Sec")

        result.update({
            'res_code': STATUS_200,
            'status_str': STR_200,
            'result': 'Done'
        })

        return result

    except Exception as es:
        print(f"An error occurred in rbi__ngo_api2: {es}")

        result.update({
            'res_code': STATUS_500,
            'status_str': str(es),
            'result': None
        })

        return result