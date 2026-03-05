
import os
import sys
import time
import argparse
import asyncio
from typing import Optional

import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from langchain_google_genai import ChatGoogleGenerativeAI


# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

# Column names — must match what your main pipeline outputs
COL_SECTION   = "Section Number"
COL_CHAPTER   = "Section Chapter Header"
COL_REG_TEXT  = "Regulatory Text BreakDown"
COL_PAGE      = "Page Number"
COL_OBL       = "Obligations"

# Colours for the output Excel
CLR_HEADER    = "1F4E79"   # dark blue header
CLR_NO_CHANGE = "E2EFDA"   # light green  — obligations identical
CLR_CHANGED   = "FFF2CC"   # light yellow — obligation wording changed
CLR_NEW       = "FCE4D6"   # light red    — section only in PDF2
CLR_REMOVED   = "D9D9D9"   # grey         — section only in PDF1
CLR_ERROR     = "F4CCCC"   # pink         — LLM call failed

NO_OBL_MARKERS = [
    "no obligations",
    "no explicit obligations",
    "no actions",
    "error: could not",
    "nan",
    "",
]



def normalise_obl(text: str) -> str:
    """
    Lowercase, strip, and remove surface differences that do NOT change
    compliance meaning, so we avoid wasting LLM calls on trivially different
    obligation text:
      - pluralisation  (entities → entity, policies → policy, etc.)
      - articles       (the / a / an)
      - trailing punctuation
      - extra whitespace
    Uses word-by-word replacement so partial words are never corrupted.
    """
    import re as _re
    if not text or str(text).lower().strip() in NO_OBL_MARKERS:
        return ""
    t = str(text).lower()
    t = " ".join(t.split())
    t = _re.sub(r"[.;,]+$", "", t).strip()

    # Word-level plural normalisation — only replaces whole words
    plural_map = {
        "entities": "entity", "organisations": "organisation",
        "organizations": "organisation", "requirements": "requirement",
        "controls": "control", "systems": "system",
        "policies": "policy", "procedures": "procedure",
        "measures": "measure", "assessments": "assessment",
        "processes": "process", "strategies": "strategy",
        "activities": "activity", "authorities": "authority",
        "parties": "party", "risks": "risk", "threats": "threat",
        "incidents": "incident", "assets": "asset", "services": "service",
        "responsibilities": "responsibility", "standards": "standard",
        "obligations": "obligation",
    }
    words = t.split()
    t = " ".join(plural_map.get(w, w) for w in words)

    # Remove leading/trailing articles and inline articles
    t = _re.sub(r" (the|a|an) ", " ", t)
    t = _re.sub(r"^(the|a|an) ", "", t)
    return " ".join(t.split())


def has_obligations(text: str) -> bool:
    return bool(normalise_obl(text))


def obligations_identical(t1: str, t2: str) -> bool:
    """
    Returns True if obligations are substantively the same after
    normalising plurals, articles, and punctuation.
    Uses fuzzy ratio as a final safety net for near-identical texts.
    """
    n1 = normalise_obl(t1)
    n2 = normalise_obl(t2)
    if n1 == n2:
        return True
    try:
        from difflib import SequenceMatcher
        ratio = SequenceMatcher(None, n1, n2).ratio()
        return ratio >= 0.97
    except Exception:
        return False


GAP_API_KEYS = [
    "", 
    "",  # key 2 
    "",  # key 3
]
GAP_API_KEYS = [k for k in GAP_API_KEYS if k.strip()]

_gap_key_index = 0


def _rotate_key():
    global _gap_key_index
    _gap_key_index = (_gap_key_index + 1) % len(GAP_API_KEYS)
    return GAP_API_KEYS[_gap_key_index]


def _make_llm(api_key: str):
    return ChatGoogleGenerativeAI(
        model="gemini-2.5-flash-lite",
        temperature=0.0,
        max_output_tokens=2048,
        google_api_key=api_key,
        max_retries=0,   # we handle retries ourselves via key rotation
    )


# ─────────────────────────────────────────────
# GAP ANALYSIS PROMPT
# ─────────────────────────────────────────────

SYSTEM_PROMPT = (
    "You are a regulatory compliance expert comparing two versions of the "
    "UAE Information Assurance Regulation. Your job is to identify gaps — "
    "obligations that appear in Version 2 but are absent or meaningfully "
    "changed compared to Version 1."
)

GAP_INSTRUCTION = (
    "Compare the two sets of compliance obligations below for the same "
    "regulatory section and identify genuine compliance gaps.\n\n"
    "VERSION 1 OBLIGATIONS:\n{obl1}\n\n"
    "VERSION 2 OBLIGATIONS:\n{obl2}\n\n"
    "Instructions:\n"
    "1. IGNORE the following — these are NOT gaps:\n"
    "   - Pluralisation only: 'entity' vs 'entities', 'organisation' vs "
    "     'organisations', 'system' vs 'systems', etc.\n"
    "   - Addition or removal of articles: 'a', 'an', 'the'\n"
    "   - Minor punctuation differences: trailing full stops, commas, semicolons\n"
    "   - Trivial rephrasing where the compliance obligation and its scope "
    "     remain exactly the same (e.g. 'shall implement' vs 'must implement')\n"
    "   - Reordering of bullet points without changing their content\n"
    "   - Capitalisation differences\n"
    "2. REPORT as a gap ONLY if the obligation in Version 2:\n"
    "   a) [NEW] — introduces a completely new compliance requirement with no "
    "      equivalent in Version 1 (same topic in different words does NOT count)\n"
    "   b) [CHANGED] — materially changes the compliance burden: different scope, "
    "      new threshold, added condition, removed exception, different timeframe, "
    "      different responsible party, or stronger/weaker obligation strength\n"
    "3. For each genuine gap, write one bullet in this exact format:\n"
    "   • [NEW] or [CHANGED] — <exact obligation text from Version 2> |"
    "     Reason: <one sentence describing what specifically changed or was added "
    "     and why it changes the compliance burden>\n"
    "4. If there are NO genuine gaps (only trivial surface differences), "
    "   respond with exactly: No gaps identified.\n"
    "5. If Version 2 has NO obligations but Version 1 did, respond with exactly: "
    "   Obligation removed in Version 2.\n"
    "6. No preamble, summary, commentary, or extra text — bullet list only."
)


# ─────────────────────────────────────────────
# ASYNC GAP ANALYSIS
# ─────────────────────────────────────────────

async def analyse_gaps_async(rows: list[dict], api_key: str) -> list[str]:
    """
    Key-rotating gap analysis. LLM created once, rotates to next key on 429.
    api_key param kept for backward compatibility but GAP_API_KEYS list is used.
    """
    global _gap_key_index

    # Seed the pool with the passed-in key if pool is empty
    if not GAP_API_KEYS and api_key:
        GAP_API_KEYS.append(api_key)

    current_key = GAP_API_KEYS[_gap_key_index]
    llm = _make_llm(current_key)
    print(f"  [Key pool] Starting gap analysis with key index {_gap_key_index} ({current_key[:8]}...)")

    results = []

    for i, row in enumerate(tqdm(rows, desc="Analysing gaps")):
        obl1 = row["obl1"] or "No obligations found in Version 1."
        obl2 = row["obl2"] or "No obligations found in Version 2."

        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": GAP_INSTRUCTION.format(
                obl1=obl1, obl2=obl2)},
        ]

        success = False
        keys_tried = 0
        while keys_tried < len(GAP_API_KEYS):
            try:
                response = await llm.ainvoke(messages)
                results.append(response.content.strip())
                success = True
                break
            except Exception as e:
                err = str(e)
                is_quota = "429" in err or "quota" in err.lower() or "ResourceExhausted" in err
                if is_quota:
                    keys_tried += 1
                    if len(GAP_API_KEYS) > 1 and keys_tried < len(GAP_API_KEYS):
                        new_key = _rotate_key()
                        print(f"  [Key pool] 429 → rotating to key {_gap_key_index} ({new_key[:8]}...)")
                        llm = _make_llm(new_key)
                        await asyncio.sleep(2)
                    else:
                        print(f"  [Single key] 429 — waiting 65s for rate window reset...")
                        await asyncio.sleep(65)
                        keys_tried = 0
                        break
                else:
                    print(f"  Non-quota error on row {i}: {err[:100]}")
                    break

        if not success:
            results.append("Error: Could not process this section.")

        await asyncio.sleep(3)   # pace: safe for free tier (20 RPM)

    return results


# ─────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────

def write_output_excel(output_rows: list[dict], output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from collections import Counter
    try:
        from excel_styling import (
            style_header_row, style_data_row, style_summary_sheet,
            freeze_and_filter, set_col_widths, status_fill, header_font,
            body_font, thin_border, top_align, center_align,
            CLR_HEADER_BG, CLR_ROW_ALT, CLR_ROW_NORM
        )
        has_styling = True
    except ImportError:
        has_styling = False

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Analysis"
    ws.sheet_view.showGridLines = False

    headers = [
        "Section Number",
        "Page (PDF1)",
        "Regulatory Text (PDF1)",
        "Obligations (PDF1)",
        "Page (PDF2)",
        "Regulatory Text (PDF2)",
        "Obligations (PDF2)",
        "Gap Analysis",
        "Status",
    ]
    col_widths = [18, 8, 42, 42, 8, 42, 42, 48, 14]
    n_cols = len(headers)

    # Write headers
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    if has_styling:
        style_header_row(ws, 1, n_cols)
        set_col_widths(ws, col_widths)
    else:
        from openpyxl.styles import Font, Alignment
        for col_idx in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = Font(bold=True, name="Calibri", size=10)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data rows
    for row_num, row in enumerate(output_rows, start=2):
        values = [
            row.get("section",   ""),
            row.get("page1",     ""),
            row.get("reg_text1", ""),
            row.get("obl1",      ""),
            row.get("page2",     ""),
            row.get("reg_text2", ""),
            row.get("obl2",      ""),
            row.get("gap",       ""),
            row.get("status",    ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_num, column=col_idx, value=val)

        if has_styling:
            status = row.get("status", "")
            fill   = status_fill(status)
            style_data_row(ws, row_num, n_cols, height=80, fill=fill)
            # Section number column — bold, centered
            sec_cell = ws.cell(row=row_num, column=1)
            sec_cell.font      = body_font(size=9, bold=True)
            sec_cell.alignment = center_align()
            # Page columns — centered
            for pg_col in [2, 5]:
                c = ws.cell(row=row_num, column=pg_col)
                c.alignment = center_align()
            # Status column — bold, centered
            st_cell = ws.cell(row=row_num, column=n_cols)
            st_cell.font      = body_font(size=9, bold=True)
            st_cell.alignment = center_align()

    if has_styling:
        freeze_and_filter(ws, "B2")

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    counts = Counter(r["status"] for r in output_rows)

    ws2["A1"] = "Gap Analysis Summary"
    ws2["A3"] = "Status"
    ws2["B3"] = "Count"

    for i, (status, count) in enumerate(counts.items(), start=4):
        ws2.cell(row=i, column=1, value=status)
        ws2.cell(row=i, column=2, value=count)

    if has_styling:
        style_summary_sheet(ws2)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\nOutput saved → {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def run(excel1_path: str, excel2_path: str,
        output_path: str, api_key: str):

    print(f"\nReading PDF1 Excel: {excel1_path}")
    df1 = pd.read_excel(excel1_path, dtype=str).fillna("")

    print(f"Reading PDF2 Excel: {excel2_path}")
    df2 = pd.read_excel(excel2_path, dtype=str).fillna("")

    # Normalise section keys for matching
    df1["_key"] = df1[COL_SECTION].str.strip().str.lower()
    df2["_key"] = df2[COL_SECTION].str.strip().str.lower()

    keys1 = set(df1["_key"])
    keys2 = set(df2["_key"])

    all_keys    = sorted(keys1 | keys2,
                         key=lambda k: df1.loc[df1["_key"]==k].index[0]
                         if k in keys1 else 9999)

    print(f"\nPDF1 sections: {len(keys1)}")
    print(f"PDF2 sections: {len(keys2)}")
    print(f"Sections only in PDF1 (removed): {len(keys1 - keys2)}")
    print(f"Sections only in PDF2 (new):     {len(keys2 - keys1)}")
    print(f"Common sections:                  {len(keys1 & keys2)}")

    # ── Build rows ──
    output_rows   = []
    needs_llm     = []   # rows that need LLM gap analysis
    needs_llm_idx = []   # their index in output_rows

    for key in all_keys:
        r1 = df1[df1["_key"] == key].iloc[0].to_dict() if key in keys1 else {}
        r2 = df2[df2["_key"] == key].iloc[0].to_dict() if key in keys2 else {}

        section   = (r1 or r2).get(COL_SECTION, key)
        chapter   = (r1 or r2).get(COL_CHAPTER, "")
        page1     = r1.get(COL_PAGE, "")
        page2     = r2.get(COL_PAGE, "")
        reg_text1 = r1.get(COL_REG_TEXT, "")
        reg_text2 = r2.get(COL_REG_TEXT, "")
        obl1      = r1.get(COL_OBL, "")
        obl2      = r2.get(COL_OBL, "")

        row = {
            "section":    section,
            "chapter":    chapter,
            "page1":      page1,
            "reg_text1":  reg_text1,
            "page2":      page2,
            "reg_text2":  reg_text2,
            "obl1":       obl1,
            "obl2":       obl2,
            "gap":        "",
            "status":     "",
        }

        if key not in keys2:
            # Section exists in PDF1 only — removed
            row["gap"]    = "This section does not appear in PDF2."
            row["status"] = "Removed"

        elif key not in keys1:
            # Section exists in PDF2 only — entirely new
            row["gap"]    = "This section is new in PDF2."
            row["status"] = "New Section"
            if has_obligations(obl2):
                # Still need LLM to describe what the new obligations are
                needs_llm.append({"obl1": "", "obl2": obl2})
                needs_llm_idx.append(len(output_rows))

        elif obligations_identical(obl1, obl2):
            # Obligations text is the same — no gap
            row["gap"]    = "No gaps identified."
            row["status"] = "No Change"

        elif not has_obligations(obl1) and not has_obligations(obl2):
            # Neither version has obligations — no gap
            row["gap"]    = "No obligations in either version."
            row["status"] = "No Change"

        else:
            # Obligations differ — needs LLM analysis
            needs_llm.append({"obl1": obl1, "obl2": obl2})
            needs_llm_idx.append(len(output_rows))

        output_rows.append(row)

    # ── Run LLM on rows that need it ──
    if needs_llm:
        print(f"\n{len(needs_llm)} sections need LLM gap analysis...")
        gap_results = asyncio.run(analyse_gaps_async(needs_llm, api_key))

        for list_idx, row_idx in enumerate(needs_llm_idx):
            gap_text = gap_results[list_idx]
            output_rows[row_idx]["gap"] = gap_text

            # Determine status from LLM response
            lower = gap_text.lower()
            if "error:" in lower:
                output_rows[row_idx]["status"] = "Error"
            elif "no gaps identified" in lower or "obligation removed" in lower:
                output_rows[row_idx]["status"] = "No Change"
            else:
                output_rows[row_idx]["status"] = "Changed"
    else:
        print("\nNo sections needed LLM analysis (all identical or no obligations).")

    # ── Write output ──
    write_output_excel(output_rows, output_path)

    # Print summary
    from collections import Counter
    counts = Counter(r["status"] for r in output_rows)
    print("\n── Summary ──")
    for status, count in counts.most_common():
        print(f"  {status:<15} {count}")


# ENTRY POINT
if __name__ == "__main__":

    GAP_API_KEYS.clear()
    GAP_API_KEYS.extend([    
        "AIzaSyA9dL96Gq5UN85AlBqKwFewyz1OvCn7RRY",
        "AIzaSyAI8sEYix7YadfmqvxJTb0hYCWcLwZm7Dc",
        "AIzaSyBwhUU82EAWQUqKMC7QGfmc6VEsKozMDPM"   
    ])

    EXCEL1_PATH = r"output\UAE IA Regulation v11\UAE IA Regulation v11.xlsx"
    EXCEL2_PATH = r"output\Modified_UAE_Circular\Modified_UAE_Circular.xlsx"
    OUTPUT_PATH = r"output\Gap_Analysis.xlsx"

    if not any(k.strip() and not k.startswith("AIzaYOUR") for k in GAP_API_KEYS):
        print("ERROR: No real API keys set. Paste your key in the entry point.")
        sys.exit(1)

    run(
        excel1_path=EXCEL1_PATH,
        excel2_path=EXCEL2_PATH,
        output_path=OUTPUT_PATH,
        api_key=GAP_API_KEYS[0],
    )