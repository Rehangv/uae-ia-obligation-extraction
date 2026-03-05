
import os
import re
import json
import pandas as pd
import tabula
import fitz
from PyPDF2 import PdfReader
from difflib import SequenceMatcher

from openpyxl import load_workbook

from rbi_constants import *
from status_code import *


# ── Constants ─────────────────────────────────────────────────────────────────
LOCAL_MOUNT_PATH = r"C:\Users\Rehan\Downloads\obligation code\src"

TOC_SECTION_THRESHOLD = 3
TABLE_MARGIN = 2

log_file = "log.txt"
_log_fh  = open(log_file, "w", encoding="utf-8")

def write_to_file(text):
    _log_fh.write(text + "\n")


# ── Utility ───────────────────────────────────────────────────────────────────

def percentage_similarity(a: str, b: str, cap: int = 2000) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a[:cap], b[:cap]).ratio() * 100


def flags_decomposer(flags):
    """Return a human-readable string of fitz font flags."""
    l = []
    if flags & 2 ** 0: l.append("superscript")
    if flags & 2 ** 1: l.append("italic")
    if flags & 2 ** 2: l.append("serifed")
    else:              l.append("sans")
    if flags & 2 ** 3: l.append("monospaced")
    else:              l.append("proportional")
    if flags & 2 ** 4: l.append("bold")
    return ", ".join(l)


# ── TOC / Cover page detector ─────────────────────────────────────────────────

def _is_cover_or_toc_page(lines_with_style, plain_lines):
    """
    Returns True when a page should be completely skipped.

    Signal A – TOC pattern:
        Any line that both starts with a heading prefix (Chapter / X.Y / M1 /
        Annex A) AND ends with a bare 1-3 digit number is a TOC line.
        Three or more such lines → TOC page.

    Signal B – Cover/title page pattern:
        Six or fewer non-empty lines, all rendered at >= 18pt font.
    """
    toc_line_count = 0
    for line in plain_lines:
        stripped = line.strip()
        if not stripped:
            continue
        has_heading_prefix = bool(
            re.match(r'^\d+\.\d+', stripped) or
            re.match(r'^(M[1-6]|T[1-9])\b', stripped) or
            re.match(r'^Chapter\s+\d+', stripped, re.I) or
            re.match(r'^Annex\s+[A-Z]\b', stripped, re.I)
        )
        has_trailing_page = bool(re.search(r'\b\d{1,3}\s*$', stripped))
        if has_heading_prefix and has_trailing_page:
            toc_line_count += 1

    if toc_line_count >= TOC_SECTION_THRESHOLD:
        return True

    non_empty = [(txt, sz) for txt, sz in lines_with_style if txt.strip()]
    if not non_empty:
        return True                             # blank page
    if len(non_empty) <= 6 and all(sz >= 18 for _, sz in non_empty):
        return True                             # cover / title page

    return False


# ── Core text extraction ──────────────────────────────────────────────────────

def extract_text_with_metadata(doc):
    """
    Walk every page of `doc` (a fitz.Document) and return a list of dicts:
        {chapter, section, family, annex, page, text, tables}

    KEY DESIGN:
    - Heading text goes into the heading key (chapter/section/family/annex)
      only — NOT into the `text` field.  This keeps Regulatory Text BreakDown
      as pure body content.
    - Bare headings with no following body text still produce a row:
      the heading label is used as fallback text.
    - Cover and TOC pages are auto-detected and skipped.
    - All heading acceptance is guarded by structural ordering
      (no stray page numbers promoted to headings).
    """
    results = []

    current = {"chapter": None, "section": None, "family": None, "annex": None}
    buffer_text = ""
    buffer_page = None

    any_chapter_seen = False
    any_section_seen = False
    any_family_seen  = False

    chapter_pattern = re.compile(r'^Chapter\s+\d+\b', re.I)
    section_pattern = re.compile(r'^(\d+\.\d+)(?!\.\d)\b')   # X.Y but NOT X.Y.Z
    annex_pattern   = re.compile(r'^Annex\s+[A-Z]\b', re.I)
    family_pattern  = re.compile(r'^(M[1-6]|T[1-9])\b')

    def flush():
        nonlocal buffer_text, buffer_page
        if buffer_text.strip():
            results.append({
                "chapter": current["chapter"],
                "section": current["section"],
                "family":  current["family"],
                "annex":   current["annex"],
                "page":    buffer_page,
                "text":    buffer_text.strip(),
                "tables":  [],
            })
        buffer_text = ""
        buffer_page = None

    pending_heading_page = None

    def flush_with_heading_fallback():
        """Flush; if buffer is empty but we have a heading emit a fallback row."""
        nonlocal buffer_text, buffer_page, pending_heading_page
        if buffer_text.strip():
            results.append({
                "chapter": current["chapter"],
                "section": current["section"],
                "family":  current["family"],
                "annex":   current["annex"],
                "page":    buffer_page,
                "text":    buffer_text.strip(),
                "tables":  [],
            })
        elif pending_heading_page is not None:
            heading_label = (
                current["annex"] or current["family"] or
                current["section"] or current["chapter"] or ""
            )
            if heading_label:
                results.append({
                    "chapter": current["chapter"],
                    "section": current["section"],
                    "family":  current["family"],
                    "annex":   current["annex"],
                    "page":    pending_heading_page,
                    "text":    heading_label,
                    "tables":  [],
                })
        buffer_text           = ""
        buffer_page           = None
        pending_heading_page  = None

    for page_number in range(doc.page_count):
        page   = doc.load_page(page_number)
        tables = page.find_tables()

        table_rects = [
            fitz.Rect(
                t.bbox[0] - TABLE_MARGIN,
                t.bbox[1] - TABLE_MARGIN,
                t.bbox[2] + TABLE_MARGIN,
                t.bbox[3] + TABLE_MARGIN,
            )
            for t in tables
        ]

        page_dict     = page.get_text("dict")
        sorted_blocks = sorted(page_dict.get("blocks", []), key=lambda b: b["bbox"][1])

        lines_with_style = []
        for block in sorted_blocks:
            if "lines" not in block:
                continue
            block_rect = fitz.Rect(block["bbox"])
            if any(block_rect.intersects(t) for t in table_rects):
                continue
            for line in block["lines"]:
                line_text = ""
                max_size  = 0
                for span in line["spans"]:
                    line_text += span["text"]
                    max_size   = max(max_size, span["size"])
                lines_with_style.append((line_text.strip(), max_size))

        plain_lines = [l[0] for l in lines_with_style]

        if _is_cover_or_toc_page(lines_with_style, plain_lines):
            buffer_text          = ""
            buffer_page          = None
            pending_heading_page = None
            continue

        for line, font_size in lines_with_style:
            if not line:
                continue
            # Skip bare page numbers
            if re.fullmatch(r'\d{1,3}', line):
                continue
            # Skip running header
            if "UAE Information Assurance Regulation" in line:
                continue

            # ── ANNEX ────────────────────────────────────────────────────────
            if annex_pattern.match(line) and font_size > 13:
                if any_chapter_seen or any_section_seen or any_family_seen:
                    candidate    = line.strip()
                    is_duplicate = current["annex"] == candidate and not buffer_text.strip()
                    if is_duplicate:
                        continue
                    flush_with_heading_fallback()
                    current              = dict.fromkeys(current, None)
                    current["annex"]     = candidate
                    pending_heading_page = page_number + 1
                    buffer_page          = page_number + 1
                    buffer_text          = ""
                    continue
                if buffer_page is None:
                    buffer_page = page_number + 1
                buffer_text += line + " "
                continue

            # ── CHAPTER ──────────────────────────────────────────────────────
            if chapter_pattern.match(line) and font_size > 15:
                flush_with_heading_fallback()
                current              = dict.fromkeys(current, None)
                current["chapter"]   = line.strip()
                pending_heading_page = page_number + 1
                buffer_page          = page_number + 1
                buffer_text          = ""
                any_chapter_seen     = True
                continue

            # ── SECTION (X.Y) ─────────────────────────────────────────────
            sec_match = section_pattern.match(line)
            if sec_match and font_size > 13:
                if re.match(r'^[MT]\d', line):
                    if buffer_page is None:
                        buffer_page = page_number + 1
                    buffer_text += line + " "
                    continue
                if not any_chapter_seen:
                    if buffer_page is None:
                        buffer_page = page_number + 1
                    buffer_text += line + " "
                    continue
                flush_with_heading_fallback()
                current["section"]   = sec_match.group(1)
                current["family"]    = None
                current["annex"]     = None
                pending_heading_page = page_number + 1
                buffer_page          = page_number + 1
                buffer_text          = ""
                any_section_seen     = True
                continue

            # ── FAMILY (M1-M6 / T1-T9) ───────────────────────────────────
            fam_match = family_pattern.match(line)
            if fam_match and font_size > 13:
                if not (any_chapter_seen or any_section_seen):
                    if buffer_page is None:
                        buffer_page = page_number + 1
                    buffer_text += line + " "
                    continue
                candidate    = fam_match.group(1)
                is_duplicate = current["family"] == candidate and not buffer_text.strip()
                if is_duplicate:
                    continue
                flush_with_heading_fallback()
                current["family"]    = candidate
                current["annex"]     = None
                pending_heading_page = page_number + 1
                buffer_page          = page_number + 1
                buffer_text          = ""
                any_family_seen      = True
                continue

            # Skip control table metadata that leaks outside table bounding boxes
            if re.fullmatch(r'P[1-4]', line.strip()):
                continue
            if line.strip() in ('Priority', 'Applicability'):
                continue
            # Skip wrapped heading continuation orphans (e.g. "and Maintenance"
            # on its own line right after a new heading with empty buffer)
            if (not buffer_text.strip() and
                    re.match(r'^(and|or|of)\s+[A-Z]', line.strip()) and
                    len(line.strip()) < 60 and
                    not re.search(r'shall|must|required', line, re.I)):
                continue

            # ── Regular paragraph text ────────────────────────────────────
            if buffer_page is None:
                buffer_page = page_number + 1
            pending_heading_page = None
            buffer_text += line + " "

    flush_with_heading_fallback()
    return results


# ── Footer page number map ────────────────────────────────────────────────────

def build_footer_page_map(reader):
    """
    Build {pdf_page_index_1based: display_page_number} from footer text.
    Uses sequential plausibility validation to reject false positives
    (e.g. numbers that appear in table cells rather than footers).
    """
    footer_page_map = {}

    for i, page in enumerate(reader.pages):
        text       = page.extract_text() or ""
        lines      = [l.strip() for l in text.split("\n") if l.strip()]
        last_lines = lines[-12:]

        found_page = None
        for line in reversed(last_lines):
            if re.fullmatch(r'\d{1,3}', line):
                candidate = int(line)
                # Must be within ±30 of physical page index to be a real footer
                if abs(candidate - (i + 1)) <= 30:
                    found_page = candidate
                    break

        footer_page_map[i + 1] = found_page if found_page else (i + 1)

    return footer_page_map


# ── Table JSON helper ─────────────────────────────────────────────────────────

def get_tab_obl_from_json(page_number, js_data):
    """
    Return readable pipe-delimited text for all tabula-extracted tables
    on `page_number`.

    tabula JSON structure:
        [{"page": N, "table_index": i, "columns": [...], "rows": [...]}, ...]
    """
    if not js_data or page_number is None:
        return ""
    table_texts = []
    for table in js_data:
        if table.get("page") == page_number:
            cols      = table.get("columns", [])
            rows      = table.get("rows",    [])
            header    = " | ".join(str(c) for c in cols)
            row_lines = [" | ".join(str(cell) for cell in row) for row in rows]
            table_texts.append(header + "\n" + "\n".join(row_lines))
    return "\n\n".join(table_texts)


# ── Applicability extraction ──────────────────────────────────────────────────

def paragraph_start_end_word(start_word, end_word, raw_text):
    """Find first matching start/end keyword pair in raw_text."""
    if not raw_text:
        return '', ''
    st_word = ''
    nd_word = ''
    try:
        for each in start_word:
            if each in raw_text:
                st_word = each
                for end in end_word:
                    if end in raw_text:
                        nd_word = end
                        break
                break
        return st_word, nd_word
    except Exception as e:
        write_to_file(f"Exception in paragraph_start_end_word: {str(e)}")
        return '', ''


def applicability_extraction(cir_date, df, path):
    """
    Extract the Applicability section from a UAE IA Regulation PDF.

    Strategy:
    1. Detect UAE PDF by scanning first 3 pages for known keywords.
    2. If UAE → run targeted applicability scanner across first 25 pages,
       scoring chunks and returning the best match.
    3. Fallback → generic keyword-boundary approach on page 0.
    """

    def _clean(txt: str) -> str:
        if not txt:
            return ""
        txt = re.sub(r'[\u0900-\u097F]+', '', txt)   # remove Hindi
        txt = re.sub(r"\(Updated.*?\)", "", txt)
        txt = txt.replace("\n", " ")
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    def _safe_start_end(start_list, end_list, text):
        try:
            tmp = paragraph_start_end_word(start_list, end_list, text)
            if tmp and isinstance(tmp, (list, tuple)) and len(tmp) == 2:
                return tmp[0] or "", tmp[1] or ""
            return "", ""
        except Exception:
            return "", ""

    def _looks_like_uae(doc):
        try:
            sample = ""
            for i in range(min(3, len(doc))):
                sample += " " + (doc[i].get_text() or "")
            s = sample.lower()
            return any(k in s for k in [
                "information assurance regulation",
                "uae ia regulation",
                "telecommunications regulatory authority",
                "implementation guidance (for information purpose only)"
            ])
        except Exception:
            return False

    def _extract_uae_applicability(doc, max_pages=25) -> str:
        """
        Scan first pages; pick the best-scoring chunk for an Applicability section.
        Penalises TOC/annex pages; rewards chunks with UAE-specific phrases.
        """
        start_keys = [
            "applicability of",
            "applicability",
            "this regulation applies",
            "this regulation shall apply",
            "scope",
            "critical entities"
        ]
        end_keys = [
            "implementation guidance (for information purpose only)",
            "implementation guidance",
            "requirements",
            "controls",
            "objective",
            "definitions",
            "annex",
            "terms and definitions",
            "bibliography"
        ]

        best_txt   = ""
        best_score = -10

        for pno in range(min(max_pages, len(doc))):
            page_txt   = doc[pno].get_text() or ""
            page_txt   = re.sub(r'[\u0900-\u097F]+', '', page_txt)
            page_lower = page_txt.lower()

            start_idx = -1
            for k in start_keys:
                i = page_lower.find(k)
                if i != -1:
                    start_idx = i
                    break
            if start_idx == -1:
                continue

            end_idx = -1
            for k in end_keys:
                i = page_lower.find(k, start_idx + 10)
                if i != -1:
                    end_idx = i
                    break

            chunk = (page_txt[start_idx:end_idx]
                     if (end_idx != -1 and end_idx > start_idx)
                     else page_txt[start_idx:])
            chunk = _clean(chunk)

            score = 0
            if "applicability of"   in chunk.lower(): score += 6
            if "this regulation"    in chunk.lower(): score += 2
            if "chapter"            in chunk.lower(): score -= 2   # TOC noise
            if "table"              in chunk.lower(): score -= 3
            if "annex"              in chunk.lower(): score -= 4
            if len(chunk) > 250:                      score += 2
            if len(chunk) > 800:                      score += 1

            if score > best_score:
                best_score = score
                best_txt   = chunk

        return best_txt

    try:
        print("APPLICABILITY CALLING------------------------------------")

        doc = fitz.open(path)

        if _looks_like_uae(doc):
            print("UAE PDF detected — running UAE applicability scan...")
            uae_txt = _extract_uae_applicability(doc, max_pages=25)
            if uae_txt and len(uae_txt) > 80:
                print("APPLICABILITY EXTRACTED (UAE):", uae_txt[:160], "...")
                return uae_txt
            print("UAE applicability not found confidently — falling back to generic logic...")

        # Generic fallback: keyword-boundary search on page 0
        start_word_appl = ["Scope", "Applicability", "This Regulation", "Critical entities"]
        end_word_appl   = ["Implementation Guidance", "Requirements", "Control", "Objective"]
        start_word_cir  = ["Information Assurance Regulation"]
        end_word_cir    = ["Annex", "Annex A", "Annex B", "Annex C", "Annex D",
                           "Annex E", "Annex F", "Annex G", "Bibliography"]

        page0 = doc[0].get_text() or ""
        page0 = re.sub(r'[\u0900-\u097F]+', '', page0)

        _safe_start_end(start_word_cir, end_word_cir, page0)   # warm-up / unused result

        page_lower = page0.lower()
        s_idx = -1
        for k in ["scope", "applicability", "this regulation applies", "critical entities"]:
            i = page_lower.find(k)
            if i != -1:
                s_idx = i
                break

        e_idx = -1
        if s_idx != -1:
            for k in ["implementation guidance", "requirements", "controls",
                      "objective", "definitions", "annex"]:
                i = page_lower.find(k, s_idx + 5)
                if i != -1:
                    e_idx = i
                    break

        if s_idx != -1:
            applicability = (page0[s_idx:e_idx]
                             if (e_idx != -1 and e_idx > s_idx)
                             else page0[s_idx:])
            applicability = _clean(applicability)
            if applicability:
                return applicability

        return "Applicability not found"

    except Exception as e:
        return f"Applicability extraction failed: {str(e)}"


# ── Main entry point ──────────────────────────────────────────────────────────

def main_rbi(path, output_directory):

    print("main_rbi inside")

    try:
        result = {}

        doc    = fitz.open(path)
        reader = PdfReader(path)

        # ── Footer page map ───────────────────────────────────────────────
        footer_page_map = build_footer_page_map(reader)

        # ── Text extraction ───────────────────────────────────────────────
        extracted_sections = extract_text_with_metadata(doc)

        fixed_sections = []
        for sec in extracted_sections:
            header   = sec.get("chapter")
            text_val = sec.get("text")
            page_no  = sec.get("page")

            sec_no = (
                sec.get("annex") or
                sec.get("family") or
                sec.get("section") or
                sec.get("chapter")
            )

            real_page = footer_page_map.get(page_no, page_no)
            fixed_sections.append((header, text_val, real_page, sec_no))

        print(f"Extracted {len(fixed_sections)} sections")

        # ── Build DataFrame ───────────────────────────────────────────────
        main_df1 = pd.DataFrame(
            fixed_sections,
            columns=[
                'Section Chapter Header',
                'Regulatory Text BreakDown',
                'Page Number',
                SECTION_NUMBER,
            ]
        ).reset_index(drop=True)

        for col in RBI_COLUMN_ORDER:
            if col not in main_df1.columns:
                main_df1[col] = 'N/A'

        # ── Cover metadata (title, date, version) ────────────────────────
        def extract_cover_metadata(reader):
            """Parse regulation title, issue date and version from the cover page."""
            text  = (reader.pages[0].extract_text() or "").strip()
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            title_parts = []
            issue_date  = 'N/A'
            version     = 'N/A'
            for line in lines:
                if re.match(
                    r'^(January|February|March|April|May|June|July|August|'
                    r'September|October|November|December)\s+\d{4}$', line, re.I):
                    issue_date = line
                elif re.match(r'^Version\s+\d+[\d\.]*$', line, re.I):
                    version = line
                else:
                    title_parts.append(line)
            title = ' '.join(title_parts).strip() or 'N/A'
            return title, issue_date, version

        reg_title, issue_date, version_str = extract_cover_metadata(reader)
        print(f"Cover metadata — Title: {reg_title!r}, Date: {issue_date!r}, Version: {version_str!r}")

        if not main_df1.empty:
            main_df1[REGULATION_TITLE]    = reg_title
            main_df1[CIRCULAR_ISSUE_DATE] = issue_date
            main_df1[CIRCULAR_NUMBER]     = version_str

        # ── Table extraction → tab.json ───────────────────────────────────
        table_lst = []
        for page_num in range(len(doc)):
            try:
                tables = tabula.read_pdf(
                    path,
                    pages=page_num + 1,
                    multiple_tables=True,
                    stream=True,
                )
                if tables:
                    for idx, table in enumerate(tables):
                        table_lst.append({
                            "page":        page_num + 1,
                            "table_index": idx,
                            "columns":     table.columns.tolist(),
                            "rows":        table.fillna("").values.tolist(),
                        })
            except Exception as e:
                print(f"Table warning p{page_num + 1}: {e}")

        doc.close()

        json_fl_path = os.path.join(output_directory, "tab.json")
        if table_lst:
            with open(json_fl_path, 'w') as f:
                json.dump(table_lst, f, indent=4)

        # ── Hidden table-text column (for obligations pipeline) ───────────
        def _get_table_text(page_val):
            try:
                pg = int(page_val) if pd.notna(page_val) else None
            except (ValueError, TypeError):
                pg = None
            return get_tab_obl_from_json(pg, table_lst)

        main_df1['_table_text'] = main_df1['Page Number'].apply(_get_table_text)

        # ── Write Excel (strip internal column) ──────────────────────────
        excel_file_path = os.path.join(
            output_directory,
            f"{os.path.basename(path).split('.')[0]}.xlsx",
        )
        excel_df = main_df1.drop(columns=['_table_text'], errors='ignore')
        excel_df.to_excel(excel_file_path, index=False)

        # ── Apply professional styling ────────────────────────────────────
        try:
            from excel_styling import (
                style_header_row, style_data_row, freeze_and_filter,
                set_col_widths, body_font, center_align, top_align,
                thin_border, alt_fill
            )

            wb_s  = load_workbook(excel_file_path)
            ws_s  = wb_s.active
            ws_s.sheet_view.showGridLines = False
            n_cols = ws_s.max_column
            n_rows = ws_s.max_row

            col_widths = [28, 55, 10, 18, 22, 12, 10, 22, 55]
            while len(col_widths) < n_cols:
                col_widths.append(20)
            set_col_widths(ws_s, col_widths[:n_cols])

            style_header_row(ws_s, 1, n_cols, height=35)

            for row_num in range(2, n_rows + 1):
                style_data_row(ws_s, row_num, n_cols,
                               height=75, fill=alt_fill(row_num))
                sec = ws_s.cell(row=row_num, column=4)
                sec.font      = body_font(size=9, bold=True)
                sec.alignment = center_align()
                pg = ws_s.cell(row=row_num, column=3)
                pg.alignment  = center_align()

            freeze_and_filter(ws_s, "B2")
            wb_s.save(excel_file_path)
            print("Styling applied to Excel output.")
        except Exception as style_err:
            print(f"Styling skipped (non-critical): {style_err}")

        result.update({
            'res_code':   STATUS_200,
            'status_str': STR_200,
            'result':     excel_file_path,
        })

    except Exception as ex:
        error_message = f"Error processing file {path}: {str(ex)}"
        print(error_message)
        result.update({
            'res_code':   STATUS_500,
            'status_str': error_message,
            'result':     path,
        })

    return result